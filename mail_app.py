# -*- coding: utf-8 -*-
"""
Основной класс приложения Mail.tm
"""

import tkinter as tk
from tkinter import ttk, messagebox
import requests
import random
import string
import os
import pyperclip
import threading
import re
import time
import platform
import winsound
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from config import (
    API_URL, ACCOUNTS_FILE, EXCEL_FILE,
    STATUS_COLORS, FONT_BASE, FONT_SMALL, FONT_BOLD, FONT_TITLE
)
from themes import THEMES
from widgets import AnimatedToggle
from imap_client import IMAPClient
from sk_generator import show_sk_window


class MailApp:
    """Основной класс приложения Mail.tm"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Mail.tm — регистрация и почтовый клиент")
        self.root.geometry("1000x650")
        
        # Переменные состояния
        self.accounts_data = []
        self.last_message_ids = set()
        self.refresh_interval_ms = 5000
        
        self.current_token = None
        self.account_type = "api"  # "api" or "imap"
        self.imap_client = None
        self.mail_tm_domains = []
        
        self.is_refreshing = False
        self.auto_refresh_job = None
        self.stop_threads = False
        self.params = {"theme": "light"}
        
        # Загружаем домены mail.tm в фоне
        threading.Thread(target=self.load_mail_tm_domains, daemon=True).start()
        
        # Статус бар
        self.status_var = tk.StringVar()
        self.status_var.set("Готов к работе")
        self.status_bar = tk.Label(root, textvariable=self.status_var, bd=1, relief=tk.FLAT, anchor=tk.W, font=FONT_SMALL)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Стили
        style = ttk.Style()
        style.configure("Treeview", rowheight=25)
        available_themes = style.theme_names()
        default_design = "default" if "default" in available_themes else style.theme_use()
        try:
            style.theme_use(default_design)
        except Exception:
            default_design = style.theme_use()
        self.design_var = tk.StringVar(value=default_design)
        
        # --- Сплиттер (PanedWindow) ---
        self.paned = tk.PanedWindow(root, orient=tk.HORIZONTAL, sashwidth=4, bg="#dcdcdc")
        self.paned.pack(fill=tk.BOTH, expand=True)
        
        # --- ЛЕВАЯ ПАНЕЛЬ (аккаунты) ---
        self.left_panel = tk.Frame(self.paned, width=260, bg="#f0f0f0")
        self.paned.add(self.left_panel, minsize=200)
        
        # Заголовок левой панели (тема)
        self.left_header = tk.Frame(self.left_panel, bg="#f0f0f0")
        self.left_header.pack(fill=tk.X, padx=10, pady=(10, 0))
        
        # Тема (светлая/темная)
        self.lbl_theme = tk.Label(self.left_header, text="Тема", bg="#f0f0f0", font=FONT_SMALL)
        self.lbl_theme.pack(side=tk.LEFT)
        self.theme_toggle = AnimatedToggle(self.left_header, on_toggle=self.on_theme_toggle_click, width=40, height=20)
        self.theme_toggle.pack(side=tk.LEFT, padx=(2, 10))
        
        # Кнопка создания
        self.btn_create = tk.Button(self.left_panel, text="Создать аккаунт", bg="#2563eb", fg="white", font=FONT_BOLD, command=self.start_create_account)
        self.btn_create.pack(pady=10, padx=10, fill=tk.X)
        
        # Список аккаунтов
        self.lbl_saved = tk.Label(self.left_panel, text="Сохраненные аккаунты", bg="#f0f0f0", font=FONT_BOLD)
        self.lbl_saved.pack(pady=(10, 0))
        
        # Панель кнопок файла
        self.file_btn_frame = tk.Frame(self.left_panel, bg="#f0f0f0")
        self.file_btn_frame.pack(pady=(0, 5), padx=10, fill=tk.X)
        
        # Кнопка обновления списка
        self.btn_reload = tk.Button(self.file_btn_frame, text="Обновить", font=FONT_SMALL, command=self.load_accounts_from_file)
        self.btn_reload.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))
        
        # Кнопка открытия файла
        self.btn_open_file = tk.Button(self.file_btn_frame, text="Файл", font=FONT_SMALL, command=self.open_accounts_file)
        self.btn_open_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2, 0))
        
        # Кнопка открытия Excel файла
        self.btn_open_excel = tk.Button(self.file_btn_frame, text="Excel", font=FONT_SMALL, command=self.open_excel_file)
        self.btn_open_excel.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(2, 0))
        
        # Кнопки управления аккаунтом
        self.btn_frame = tk.Frame(self.left_panel, bg="#f0f0f0")
        self.btn_frame.pack(side=tk.BOTTOM, pady=10, fill=tk.X, padx=10)
        
        self.btn_copy_email = tk.Button(self.btn_frame, text="Email", command=self.copy_email, font=FONT_SMALL)
        self.btn_copy_email.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        
        self.btn_copy_pass = tk.Button(self.btn_frame, text="Пароль", command=self.copy_pass, font=FONT_SMALL)
        self.btn_copy_pass.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        
        # Кнопка для генерации SK данных
        self.btn_sk = tk.Button(self.btn_frame, text="SK Info", command=self._show_sk_window, font=FONT_SMALL)
        self.btn_sk.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)
        
        # Список аккаунтов
        self.acc_listbox = tk.Listbox(self.left_panel, height=20, exportselection=False)
        self.acc_listbox.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)
        self.acc_listbox.bind('<<ListboxSelect>>', self.on_account_select)
        
        # Контекстное меню для аккаунтов
        self.context_menu = tk.Menu(root, tearoff=0)
        self.context_menu.add_command(label="Статус: Не зарегистрирован", command=lambda: self.set_account_status("not_registered"))
        self.context_menu.add_command(label="Статус: Зарегистрирован", command=lambda: self.set_account_status("registered"))
        self.context_menu.add_command(label="Статус: Plus", command=lambda: self.set_account_status("plus"))
        
        self.acc_listbox.bind("<Button-3>", self.show_context_menu)
        
        # Загружаем сохраненные аккаунты сразу
        self.load_accounts_from_file()
        
        # --- ПРАВАЯ ПАНЕЛЬ (ПИСЬМА) ---
        self.right_panel = tk.Frame(self.paned)
        self.paned.add(self.right_panel, minsize=400)
        
        # Заголовок текущей почты
        self.header_frame = tk.Frame(self.right_panel, bg="#ddd")
        self.header_frame.pack(fill=tk.X)
        
        self.lbl_current_email = tk.Label(self.header_frame, text="Выберите аккаунт слева", font=FONT_TITLE, bg="#ddd", pady=10)
        self.lbl_current_email.pack(side=tk.LEFT, padx=10, fill=tk.X, expand=True)
        
        self.btn_refresh = tk.Button(self.header_frame, text="Обновить", command=self.on_manual_refresh, bg="#2196F3", fg="white", font=FONT_SMALL)
        self.btn_refresh.pack(side=tk.RIGHT, padx=10)
        
        # Кнопки статуса
        self.status_frame = tk.Frame(self.header_frame, bg="#ddd")
        self.status_frame.pack(side=tk.RIGHT, padx=5)
        
        self.btn_nr = tk.Button(self.status_frame, text="Не рег", bg="white", font=FONT_SMALL, command=lambda: self.set_account_status("not_registered"))
        self.btn_nr.pack(side=tk.LEFT, padx=2)
        
        self.btn_reg = tk.Button(self.status_frame, text="Рег", bg="#B3E5FC", font=FONT_SMALL, command=lambda: self.set_account_status("registered"))
        self.btn_reg.pack(side=tk.LEFT, padx=2)
        
        self.btn_plus = tk.Button(self.status_frame, text="Plus", bg="#80DEEA", font=FONT_SMALL, command=lambda: self.set_account_status("plus"))
        self.btn_plus.pack(side=tk.LEFT, padx=2)
        
        # Список писем (Treeview)
        columns = ("sender", "subject", "date", "msg_id")
        self.tree = ttk.Treeview(self.right_panel, columns=columns, displaycolumns=("sender", "subject", "date"), show="headings", height=8)
        self.tree.heading("sender", text="От кого")
        self.tree.heading("subject", text="Тема")
        self.tree.heading("date", text="Время")
        self.tree.column("sender", width=150)
        self.tree.column("subject", width=300)
        self.tree.column("date", width=110, anchor="center")
        self.tree.column("msg_id", width=0, stretch=False)
        self.tree.pack(fill=tk.X, padx=10, pady=10)
        self.tree.bind("<<TreeviewSelect>>", self.on_message_select)
        
        # Область просмотра письма
        self.lbl_msg_title = tk.Label(self.right_panel, text="Содержание письма:", anchor="w", font=FONT_BOLD)
        self.lbl_msg_title.pack(fill=tk.X, padx=10)
        
        # Кнопка копирования кода
        self.btn_copy_code = tk.Button(self.right_panel, text="Копировать код", bg="#FF9800", fg="white", font=FONT_BOLD)
        self.btn_copy_code.pack(fill=tk.X, padx=10, pady=5)
        self.btn_copy_code.pack_forget()
        
        self.msg_text = tk.Text(self.right_panel, wrap=tk.WORD, height=15, font=FONT_BASE)
        self.msg_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.msg_text.insert(tk.END, "Выберите письмо слева, чтобы увидеть содержимое.")
        
        print(f"[*] Используемый файл аккаунтов: {ACCOUNTS_FILE}")
        
        # Current theme
        self.set_theme("light")
        
        # Запуск цикла автообновления
        self.start_auto_refresh()
    
    def load_mail_tm_domains(self):
        """Загрузка доменов mail.tm"""
        try:
            res = requests.get(f"{API_URL}/domains")
            if res.status_code == 200:
                data = res.json()['hydra:member']
                self.mail_tm_domains = [d['domain'] for d in data]
                print(f"[*] Loaded {len(self.mail_tm_domains)} mail.tm domains")
        except:
            pass
    
    def _show_sk_window(self):
        """Открывает окно генератора данных Южной Кореи"""
        theme_name = self.params.get("theme", "light")
        show_sk_window(self.root, theme_name)
    
    def play_notification_sound(self, count=1):
        """Проигрывает звук при появлении новых писем"""
        def _beep():
            for _ in range(max(1, count)):
                try:
                    if platform.system() == "Windows":
                        winsound.MessageBeep(winsound.MB_ICONASTERISK)
                    else:
                        print("\a", end="", flush=True)
                except Exception:
                    pass
                time.sleep(0.1)
        
        threading.Thread(target=_beep, daemon=True).start()
    
    def update_status(self, text):
        """Безопасное обновление статуса из другого потока"""
        self.root.after(0, lambda: self.status_var.set(text))
    
    def open_accounts_file(self):
        """Открывает файл аккаунтов в системном редакторе"""
        try:
            if not os.path.exists(ACCOUNTS_FILE):
                with open(ACCOUNTS_FILE, "w") as f:
                    pass
            os.startfile(ACCOUNTS_FILE)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")
    
    def open_excel_file(self):
        """Открывает Excel файл аккаунтов"""
        try:
            if not os.path.exists(EXCEL_FILE):
                self.save_accounts_to_excel()
            os.startfile(EXCEL_FILE)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть Excel файл:\n{e}")
    
    def save_accounts_to_excel(self):
        """Сохраняет данные аккаунтов в Excel файл"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Аккаунты"
            
            headers = ["Логин/Пароль", "Логин", "Пароль", "Подписка"]
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            
            status_labels = {
                "not_registered": "Не зарегистрирован",
                "registered": "Зарегистрирован",
                "plus": "Plus"
            }
            
            status_fills = {
                "not_registered": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                "registered": PatternFill(start_color="B3E5FC", end_color="B3E5FC", fill_type="solid"),
                "plus": PatternFill(start_color="80DEEA", end_color="80DEEA", fill_type="solid")
            }
            
            for row, account in enumerate(self.accounts_data, 2):
                email = account.get("email", "")
                password = account.get("password", "")
                status = account.get("status", "not_registered")
                
                ws.cell(row=row, column=1, value=f"{email} / {password}")
                ws.cell(row=row, column=2, value=email)
                ws.cell(row=row, column=3, value=password)
                status_cell = ws.cell(row=row, column=4, value=status_labels.get(status, status))
                status_cell.alignment = Alignment(horizontal="center")
                
                row_fill = status_fills.get(status, status_fills["not_registered"])
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = row_fill
            
            wb.save(EXCEL_FILE)
            
        except Exception as e:
            print(f"Ошибка сохранения Excel: {e}")
    
    def load_accounts_from_file(self):
        """Загрузка аккаунтов из файла"""
        self.acc_listbox.delete(0, tk.END)
        self.accounts_data = []
        
        if os.path.exists(ACCOUNTS_FILE):
            try:
                needs_save = False
                
                with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue
                    
                    email = ""
                    password = ""
                    status = "not_registered"
                    
                    if " / " in line:
                        parts = line.split(" / ")
                        if len(parts) >= 2:
                            email = parts[0].strip()
                            password = parts[1].strip()
                            if len(parts) >= 3:
                                status = parts[2].strip()
                    elif ":" in line:
                        parts = line.split(":", 1)
                        if len(parts) == 2:
                            email, password = parts[0].strip(), parts[1].strip()
                            needs_save = True
                    
                    if email and password:
                        self.accounts_data.append({
                            "email": email,
                            "password": password,
                            "status": status
                        })
                        
                        display_text = f"{email} / {password}"
                        self.acc_listbox.insert(tk.END, display_text)
                
                if needs_save:
                    self.save_accounts_to_file()
                    self.update_status(f"Аккаунты конвертированы и загружены: {len(self.accounts_data)}")
                else:
                    self.save_accounts_to_excel()
                    self.update_status(f"Загружено аккаунтов: {len(self.accounts_data)}")
                
                self.update_listbox_colors()
                
            except Exception as e:
                messagebox.showerror("Ошибка чтения файла", str(e))
        else:
            self.update_status("Файл accounts.txt не найден")
    
    def start_create_account(self):
        """Запуск создания аккаунта"""
        self.btn_create.config(state=tk.DISABLED)
        self.update_status("Регистрация... (Подождите)")
        threading.Thread(target=self.create_account_thread, daemon=True).start()
    
    def create_account_thread(self):
        """Поток создания аккаунта"""
        try:
            domain_res = requests.get(f"{API_URL}/domains")
            if domain_res.status_code != 200:
                self.root.after(0, lambda: messagebox.showerror("Ошибка", "Не удалось получить список доменов"))
                self.root.after(0, lambda: self.btn_create.config(state=tk.NORMAL))
                return
            
            domains = domain_res.json()['hydra:member']
            domain = random.choice(domains)['domain']
            
            username = ''.join(random.choice(string.ascii_lowercase + string.digits) for _ in range(10))
            chars = string.ascii_letters + string.digits
            password = ''.join(random.choice(chars) for _ in range(12))
            
            email = f"{username}@{domain}"
            
            payload = {"address": email, "password": password}
            res = requests.post(f"{API_URL}/accounts", json=payload)
            
            if res.status_code == 201:
                self.root.after(0, lambda: self._on_account_created(email, password))
            else:
                self.update_status("Ошибка регистрации")
                self.root.after(0, lambda: messagebox.showerror("Ошибка", f"Код: {res.status_code}\n{res.text}"))
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Ошибка", str(e)))
        finally:
            self.root.after(0, lambda: self.btn_create.config(state=tk.NORMAL))
    
    def _on_account_created(self, email, password):
        """Обработка созданного аккаунта"""
        self.accounts_data.append({
            "email": email,
            "password": password,
            "status": "not_registered"
        })
        
        display_text = f"{email} / {password}"
        self.acc_listbox.insert(tk.END, display_text)
        self.update_listbox_colors()
        
        self.acc_listbox.selection_clear(0, tk.END)
        self.acc_listbox.selection_set(tk.END)
        
        self.save_accounts_to_file()
        
        self.status_var.set(f"Создан: {email}")
        self.on_account_select(None)
    
    def on_theme_toggle_click(self, is_on):
        """Обработка переключения темы"""
        self.set_theme("dark" if is_on else "light")
    
    def set_theme(self, theme_name):
        """Установка темы оформления"""
        self.params["theme"] = theme_name
        colors = THEMES[theme_name]
        accent_bg = colors.get("accent", colors["btn_bg"])
        accent_fg = colors.get("accent_fg", colors["btn_fg"])
        
        # Root
        self.root.config(bg=colors["bg"])
        self.paned.config(bg=colors["header_bg"])
        self.status_bar.config(bg=colors["status_bg"], fg=colors["status_fg"])
        
        # Left Panel Components
        self.left_panel.config(bg=colors["panel_bg"])
        self.left_header.config(bg=colors["panel_bg"])
        
        for widget in self.left_header.winfo_children():
            if isinstance(widget, tk.Label):
                widget.config(bg=colors["panel_bg"], fg=colors["fg"])
        
        self.lbl_saved.config(bg=colors["panel_bg"], fg=colors["fg"])
        self.file_btn_frame.config(bg=colors["panel_bg"])
        self.btn_frame.config(bg=colors["panel_bg"])
        
        # Buttons (Generic)
        generic_btns = [
            self.btn_reload, self.btn_open_file, self.btn_open_excel,
            self.btn_copy_email, self.btn_copy_pass, self.btn_sk
        ]
        for btn in generic_btns:
            btn.config(bg=colors["btn_bg"], fg=colors["btn_fg"], activebackground=colors["btn_bg"], activeforeground=colors["btn_fg"], relief=tk.FLAT, bd=0)
        
        # Primary buttons
        primary_btns = [self.btn_create, self.btn_refresh, self.btn_copy_code]
        for btn in primary_btns:
            btn.config(bg=accent_bg, fg=accent_fg, activebackground=accent_bg, activeforeground=accent_fg, relief=tk.FLAT, bd=0)
        
        # Listbox
        self.acc_listbox.config(bg=colors["list_bg"], fg=colors["list_fg"], selectbackground=accent_bg, selectforeground=accent_fg, relief=tk.FLAT, borderwidth=0, highlightthickness=0)
        self.update_listbox_colors()
        
        # Right Panel Components
        self.right_panel.config(bg=colors["bg"])
        self.header_frame.config(bg=colors["header_bg"])
        self.status_frame.config(bg=colors["header_bg"])
        
        self.lbl_current_email.config(bg=colors["header_bg"], fg=colors["fg"])
        self.lbl_msg_title.config(bg=colors["bg"], fg=colors["fg"])
        self.btn_refresh.config(bg=accent_bg, fg=accent_fg, activebackground=accent_bg, activeforeground=accent_fg)
        self.btn_nr.config(bg=colors["btn_bg"], fg=colors["btn_fg"])
        self.btn_reg.config(bg="#B3E5FC", fg="black")
        self.btn_plus.config(bg="#80DEEA", fg="black")
        
        # Text
        self.msg_text.config(bg=colors["text_bg"], fg=colors["text_fg"], insertbackground=colors["fg"], relief=tk.FLAT, borderwidth=1, highlightthickness=0)
        
        # Treeview Style
        style = ttk.Style()
        selected_design = self.design_var.get() if hasattr(self, 'design_var') else style.theme_use()
        if selected_design not in style.theme_names():
            selected_design = "default" if "default" in style.theme_names() else style.theme_use()
        try:
            style.theme_use(selected_design)
        except Exception:
            selected_design = style.theme_use()
        if hasattr(self, "design_var"):
            self.design_var.set(selected_design)
        
        style.configure("Treeview",
                        background=colors["list_bg"],
                        foreground=colors["list_fg"],
                        fieldbackground=colors["list_bg"],
                        rowheight=25,
                        borderwidth=0)
        style.configure("Treeview.Heading",
                        background=accent_bg,
                        foreground=accent_fg,
                        relief="flat")
        style.map("Treeview",
                  background=[("selected", accent_bg)],
                  foreground=[("selected", accent_fg)])
        
        if hasattr(self, 'lbl_theme'):
            self.lbl_theme.config(bg=colors["panel_bg"], fg=colors["fg"])
    
    def on_design_change(self, event=None):
        """Изменение дизайна (ttk theme)"""
        selected = self.design_var.get()
        style = ttk.Style()
        try:
            style.theme_use(selected)
            self.update_status(f"Дизайн изменен: {selected}")
        except Exception as e:
            self.update_status(f"Ошибка смены дизайна: {e}")
        
        self.set_theme(self.params.get("theme", "light"))
    
    def update_listbox_colors(self):
        """Обновление цветов списка аккаунтов"""
        theme = self.params.get("theme", "light")
        for i in range(self.acc_listbox.size()):
            if i < len(self.accounts_data):
                status = self.accounts_data[i].get("status", "not_registered")
                color = STATUS_COLORS.get(status, {}).get(theme, "white")
                fg_color = "#111827" if theme == "light" else "#e2e8f0"
                if status in ("registered", "plus"):
                    fg_color = "#0b1220"
                if status == "not_registered" and theme == "dark":
                    fg_color = "#e2e8f0"
                
                self.acc_listbox.itemconfig(i, {'bg': color, 'fg': fg_color})
    
    def on_account_select(self, event):
        """Выбор аккаунта"""
        selection = self.acc_listbox.curselection()
        if not selection:
            return
        
        data = self.acc_listbox.get(selection[0])
        
        if " / " in data:
            email, password = data.split(" / ", 1)
        elif ":" in data:
            email, password = data.split(":", 1)
        else:
            return
        
        self.lbl_current_email.config(text=f"Аккаунт: {email}")
        self.last_message_ids = set()
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.msg_text.delete(1.0, tk.END)
        self.msg_text.insert(tk.END, "Загрузка...")
        
        self.update_status("Авторизация...")
        threading.Thread(target=self.login_thread, args=(email, password), daemon=True).start()
    
    def login_thread(self, email_addr, password):
        """Поток авторизации"""
        domain = email_addr.split("@")[-1]
        self.current_token = None
        if self.imap_client:
            try:
                self.imap_client.logout()
            except Exception:
                pass
            self.imap_client = None
        
        is_mail_tm = domain in self.mail_tm_domains or domain.endswith("mail.tm")
        
        success = False
        
        if is_mail_tm:
            try:
                payload = {"address": email_addr, "password": password}
                res = requests.post(f"{API_URL}/token", json=payload)
                if res.status_code == 200:
                    self.current_token = res.json()['token']
                    self.account_type = "api"
                    success = True
                else:
                    print(f"API Login failed: {res.status_code}")
            except Exception as e:
                print(f"API Error: {e}")
        
        if not success:
            if self.imap_client:
                self.imap_client.logout()
            
            self.imap_client = IMAPClient(host="imap.firstmail.ltd")
            if self.imap_client.login(email_addr, password):
                self.account_type = "imap"
                success = True
            else:
                fallback_host = f"imap.{domain}"
                print(f"Trying fallback IMAP: {fallback_host}")
                self.imap_client = IMAPClient(host=fallback_host)
                if self.imap_client.login(email_addr, password):
                    self.account_type = "imap"
                    success = True
        
        if success:
            self.last_message_ids = set()
            self.update_status(f"Вход выполнен ({self.account_type.upper()}). Получаю письма...")
            self.refresh_inbox_thread()
        else:
            self.update_status("Ошибка входа (API и IMAP недоступны)")
            self.current_token = None
            self.imap_client = None
    
    def on_manual_refresh(self):
        """Ручное обновление писем"""
        self.update_status("Обновление писем...")
        threading.Thread(target=self.refresh_inbox_thread, daemon=True).start()
    
    def start_auto_refresh(self):
        """Запускает таймер автообновления"""
        if self.stop_threads:
            return
        
        should_refresh = (
            (self.account_type == "api" and self.current_token) or
            (self.account_type == "imap" and self.imap_client)
        )
        if should_refresh and not self.is_refreshing:
            threading.Thread(target=self.refresh_inbox_thread, daemon=True).start()
        
        self.root.after(self.refresh_interval_ms, self.start_auto_refresh)
    
    def refresh_inbox_thread(self):
        """Поток обновления писем"""
        if self.is_refreshing:
            return
        if self.account_type == "api" and not self.current_token:
            return
        if self.account_type == "imap" and not self.imap_client:
            return
        
        self.is_refreshing = True
        self.root.after(0, self.show_inbox_loading_state)
        self.root.after(0, self.show_loading_messages_text)
        try:
            messages = []
            if self.account_type == "api":
                headers = {"Authorization": f"Bearer {self.current_token}"}
                res = requests.get(f"{API_URL}/messages", headers=headers)
                if res.status_code == 200:
                    messages = res.json()['hydra:member']
                else:
                    self.root.after(0, lambda: self.update_status(f"Ошибка загрузки писем: {res.status_code}"))
            elif self.account_type == "imap":
                messages = self.imap_client.get_messages(limit=20)
            
            self.root.after(0, lambda: self._update_inbox_ui(messages))
        except Exception as e:
            print(f"Background update error: {e}")
        finally:
            self.is_refreshing = False
    
    def _update_inbox_ui(self, messages):
        """Обновление таблицы писем"""
        selected = self.tree.selection()
        selected_id = None
        if selected:
            values = self.tree.item(selected[0]).get("values", [])
            if len(values) >= 4:
                selected_id = values[3]
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        seen_ids = set()
        for msg in messages:
            sender = msg.get('from', {}).get('address', "Неизвестно")
            subject = msg.get('subject') or "(без темы)"
            date_raw = msg.get('createdAt') or ""
            msg_id = msg.get('id')
            try:
                dt = datetime.fromisoformat(str(date_raw).replace("Z", "+00:00"))
                date_str = dt.strftime("%H:%M:%S")
            except Exception:
                date_str = date_raw
            
            item_id = self.tree.insert("", 0, values=(sender, subject, date_str, msg_id))
            seen_ids.add(msg_id)
            
            if selected_id and msg_id == selected_id:
                self.tree.selection_set(item_id)
        
        if not messages:
            self.msg_text.delete(1.0, tk.END)
            self.msg_text.insert(tk.END, "Нет новых писем.")
        
        new_ids = [mid for mid in seen_ids if mid and mid not in self.last_message_ids]
        if self.last_message_ids and new_ids:
            self.play_notification_sound(len(new_ids))
        self.last_message_ids = seen_ids
        
        self.status_var.set(f"Обновлено: {datetime.now().strftime('%H:%M:%S')} • писем: {len(messages)}")
    
    def show_inbox_loading_state(self):
        """Показываем индикатор загрузки"""
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.tree.insert("", 0, values=("Загрузка писем...", "", "", "loading"))
        except Exception:
            pass
    
    def show_loading_messages_text(self):
        """Показываем текст загрузки"""
        try:
            if self.tree.selection():
                return
            self.btn_copy_code.pack_forget()
            self.msg_text.delete(1.0, tk.END)
            self.msg_text.insert(tk.END, "Загрузка сообщений...")
        except Exception:
            pass
    
    def on_message_select(self, event):
        """Выбор письма"""
        selected_item = self.tree.selection()
        if not selected_item:
            return
        
        item = self.tree.item(selected_item)
        values = item.get('values', [])
        if len(values) < 4:
            return
        msg_id = values[3]
        sender = values[0]
        subject = values[1]
        
        self.btn_copy_code.pack_forget()
        
        self.msg_text.delete(1.0, tk.END)
        self.msg_text.insert(tk.END, "Загрузка...")
        
        threading.Thread(target=self.load_message_thread, args=(msg_id, sender, subject), daemon=True).start()
    
    def load_message_thread(self, msg_id, sender=None, subject=None):
        """Поток загрузки письма"""
        if self.account_type == "api" and not self.current_token:
            return
        if self.account_type == "imap" and not self.imap_client:
            return
        
        try:
            if self.account_type == "api":
                headers = {"Authorization": f"Bearer {self.current_token}"}
                res = requests.get(f"{API_URL}/messages/{msg_id}", headers=headers)
                if res.status_code == 200:
                    data = res.json()
                    text = data.get('text') or data.get('html') or "Нет текстового содержимого"
                    self.root.after(0, lambda: self._show_message_content(data, text))
                else:
                    self.root.after(0, lambda: self.msg_text.insert(tk.END, f"\nОшибка загрузки письма: {res.status_code}"))
            elif self.account_type == "imap":
                text = self.imap_client.get_message_content(msg_id)
                data = {
                    "from": {"address": sender or "IMAP Sender"},
                    "subject": subject or "IMAP Message"
                }
                self.root.after(0, lambda: self._show_message_content(data, text, is_imap=True))
        except Exception as e:
            self.root.after(0, lambda: self.msg_text.insert(tk.END, f"\nError: {e}"))
    
    def _show_message_content(self, data, text, is_imap=False):
        """Отображение содержимого письма"""
        self.btn_copy_code.pack_forget()
        
        self.msg_text.delete(1.0, tk.END)
        sender = data.get('from', {}).get('address', 'Неизвестно')
        subject = data.get('subject', '(без темы)')
        self.msg_text.insert(tk.END, f"От: {sender}\n")
        self.msg_text.insert(tk.END, f"Тема: {subject}\n")
        
        self.msg_text.insert(tk.END, "-" * 50 + "\n\n")
        self.msg_text.insert(tk.END, text)
        
        match = re.search(r'\b(\d{6})\b', text)
        if match:
            code = match.group(1)
            self.btn_copy_code.config(text=f"📋 Скопировать код: {code}", command=lambda: self.copy_code_to_clipboard(code))
            self.btn_copy_code.pack(before=self.msg_text, fill=tk.X, padx=10, pady=5)
    
    def copy_code_to_clipboard(self, code):
        """Копирование кода в буфер"""
        pyperclip.copy(code)
        self.status_var.set(f"Код {code} скопирован в буфер!")
    
    def copy_email(self):
        """Копирование email"""
        selection = self.acc_listbox.curselection()
        if selection:
            data = self.acc_listbox.get(selection[0])
            if " / " in data:
                email = data.split(" / ")[0]
            elif ":" in data:
                email = data.split(":")[0]
            else:
                return
            pyperclip.copy(email)
            self.status_var.set("Email скопирован в буфер")
    
    def copy_pass(self):
        """Копирование пароля"""
        selection = self.acc_listbox.curselection()
        if selection:
            data = self.acc_listbox.get(selection[0])
            if " / " in data:
                password = data.split(" / ")[1]
            elif ":" in data:
                password = data.split(":")[1]
            else:
                return
            pyperclip.copy(password)
            self.status_var.set("Пароль скопирован в буфер")
    
    def show_context_menu(self, event):
        """Показ контекстного меню"""
        try:
            self.acc_listbox.selection_clear(0, tk.END)
            self.acc_listbox.selection_set(self.acc_listbox.nearest(event.y))
            self.acc_listbox.activate(self.acc_listbox.nearest(event.y))
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def set_account_status(self, status):
        """Установка статуса аккаунта"""
        selection = self.acc_listbox.curselection()
        if not selection:
            return
        
        idx = selection[0]
        
        if idx < len(self.accounts_data):
            self.accounts_data[idx]['status'] = status
            self.update_listbox_colors()
            self.save_accounts_to_file()
            self.update_status(f"Статус обновлен: {status}")
    
    def save_accounts_to_file(self):
        """Сохранение аккаунтов в файл"""
        try:
            with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
                for item in self.accounts_data:
                    line = f"{item['email']} / {item['password']} / {item['status']}\n"
                    f.write(line)
            self.save_accounts_to_excel()
        except Exception as e:
            messagebox.showerror("Ошибка сохранения", str(e))
