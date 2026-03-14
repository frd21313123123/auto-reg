# -*- coding: utf-8 -*-
"""
Основной класс приложения Mail.tm — Modern UI
"""

import tkinter as tk
from tkinter import ttk, messagebox
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import random
import string
import os
import sys
import pyperclip
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import time
import platform
import winsound
import ctypes
from datetime import datetime, timedelta
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .config import (
    API_URL,
    ACCOUNTS_FILE,
    EXCEL_FILE,
    STATUS_COLORS,
    FONT_BASE,
    FONT_SMALL,
    FONT_BOLD,
    FONT_TITLE,
    FONT_SECTION,
)
from .themes import THEMES
from .widgets import HoverButton, AnimatedToggle, ThemedCheckbox, SectionLabel
from .imap_client import IMAPClient
from .sk_generator import show_sk_window
from .in_generator import show_in_window
from .minesweeper import show_minesweeper
from .hotkey_settings import HotkeySettings, show_settings_window
from .live_cards_pool import show_pre_generator_window


class MailApp:
    """Основной класс приложения Mail.tm"""

    # Отступы
    PAD_X = 12
    PAD_Y = 6
    DEFAULT_THEME = "dark"
    WINDOW_FADE_STEPS = 14
    WINDOW_FADE_DELAY = 18
    THEME_TRANSITION_STEPS = 8
    THEME_TRANSITION_DELAY = 16

    def __init__(self, root):
        self.root = root
        self.root.title("Mail.tm — Auto Registration")
        self.root.geometry("1050x680")
        self.root.minsize(800, 500)
        self._window_alpha_supported = self._set_window_alpha(0.0)

        # Устанавливаем иконку окна
        try:
            icon_path = os.path.join(
                os.path.dirname(os.path.dirname(__file__)), "assets", "icon.ico"
            )
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass

        # Переменные состояния
        self.accounts_data = []
        self.last_message_ids = set()
        self.refresh_interval_ms = 5000

        self.current_token = None
        self.account_type = "api"  # "api" or "imap"
        self.imap_client = None
        self.mail_tm_domains = []

        # Сохраняем учетные данные для переподключения при смене VPN
        self.current_email = None
        self.current_password = None

        # HTTP сессия с настройками повторного подключения
        self.http_session = self._create_http_session()

        self.is_refreshing = False
        self.auto_refresh_job = None
        self.stop_threads = False
        self.params = {"theme": self.DEFAULT_THEME}
        self.is_pinned = False
        self._ban_thread_local = threading.local()
        self._ban_thread_sessions = []
        self._ban_thread_sessions_lock = threading.Lock()
        self._ban_imap_host_cache = {}
        self._ban_imap_host_lock = threading.Lock()
        self._theme_transition_job = None

        # Загружаем домены mail.tm в фоне
        threading.Thread(target=self.load_mail_tm_domains, daemon=True).start()

        # Инициализация Faker
        self.fake = Faker("en_US")

        # ======== BUILD UI ========
        self._build_ui()

        print(f"[*] Используемый файл аккаунтов: {ACCOUNTS_FILE}")

        # Применяем тему
        self.set_theme(self.DEFAULT_THEME, animate=False)

        # Загружаем аккаунты
        self.load_accounts_from_file()

        # Генерируем случайные данные
        self.generate_random_person()

        # Запуск цикла автообновления
        self.start_auto_refresh()

        # Регистрация горячих клавиш
        self._setup_hotkeys()
        self.root.after(40, self._animate_window_open)

    def _build_ui(self):
        """Построение всего интерфейса."""
        colors = THEMES[self.params.get("theme", self.DEFAULT_THEME)]

        # Основной контейнер
        self.root_container = tk.Frame(self.root, bg=colors["bg"])
        self.root_container.pack(fill=tk.BOTH, expand=True)

        # Статус бар (внизу)
        self.status_var = tk.StringVar(value="Готов к работе")
        self.status_bar = tk.Label(
            self.root_container,
            textvariable=self.status_var,
            bd=0,
            relief=tk.FLAT,
            anchor=tk.W,
            font=FONT_SMALL,
            padx=self.PAD_X,
            pady=4,
        )
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Тонкий разделитель над статус-баром
        tk.Frame(self.root_container, bg=colors["separator"], height=1).pack(
            side=tk.BOTTOM, fill=tk.X
        )

        # Стили Treeview
        style = ttk.Style()
        available_themes = style.theme_names()
        default_design = (
            "default" if "default" in available_themes else style.theme_use()
        )
        try:
            style.theme_use(default_design)
        except Exception:
            default_design = style.theme_use()
        self.design_var = tk.StringVar(value=default_design)
        style.configure("Treeview", rowheight=28)

        # Стили скроллбаров (тёмные для clam/dark)
        style.configure(
            "Dark.Vertical.TScrollbar",
            background=colors["btn_bg"],
            troughcolor=colors["panel_bg"],
            bordercolor=colors["panel_bg"],
            arrowcolor=colors["fg"],
            lightcolor=colors["panel_bg"],
            darkcolor=colors["panel_bg"],
            gripcount=0,
            borderwidth=0,
            width=12,
        )
        style.map(
            "Dark.Vertical.TScrollbar",
            background=[("active", colors["btn_hover"]), ("pressed", colors["btn_hover"])],
        )

        # --- Сплиттер ---
        self.paned = tk.PanedWindow(
            self.root_container, orient=tk.HORIZONTAL, sashwidth=2,
            bg=colors["separator"], bd=0
        )
        self.paned.pack(fill=tk.BOTH, expand=True)

        # --- ЛЕВАЯ ПАНЕЛЬ ---
        self._build_left_panel(colors)

        # --- ПРАВАЯ ПАНЕЛЬ ---
        self._build_right_panel(colors)

    def _build_left_panel(self, colors):
        """Построение левой панели (аккаунты + инструменты)."""
        self.left_panel = tk.Frame(self.paned, width=280, bg=colors["panel_bg"])
        self.paned.add(self.left_panel, minsize=240)
        self.left_panel.grid_columnconfigure(0, weight=1)

        row = 0

        # ---- HEADER: Лого + Тема ----
        self.left_header = tk.Frame(self.left_panel, bg=colors["panel_bg"])
        self.left_header.grid(row=row, column=0, sticky="ew", padx=self.PAD_X,
                              pady=(self.PAD_X, 4))
        row += 1

        self.lbl_app_title = tk.Label(
            self.left_header, text="Mail.tm", font=("Segoe UI", 16, "bold"),
            bg=colors["panel_bg"], fg=colors["accent"]
        )
        self.lbl_app_title.pack(side=tk.LEFT)

        # Тема toggle (справа в header)
        self.theme_frame = tk.Frame(self.left_header, bg=colors["panel_bg"])
        self.theme_frame.pack(side=tk.RIGHT)

        self.lbl_theme_icon = tk.Label(
            self.theme_frame, text="☀", font=("Segoe UI", 11),
            bg=colors["panel_bg"], fg=colors["muted"]
        )
        self.lbl_theme_icon.pack(side=tk.LEFT, padx=(0, 4))

        self.theme_toggle = AnimatedToggle(
            self.theme_frame, on_toggle=self.on_theme_toggle_click,
            width=44, height=22, bg_on=colors["accent"], bg_off=colors["btn_bg"]
        )
        self.theme_toggle.pack(side=tk.LEFT)

        # Pin button
        self.btn_pin = HoverButton(
            self.left_header, text="📌", font=("Segoe UI", 11),
            bg=colors["panel_bg"], fg=colors["muted"],
            hover_bg=colors["btn_hover"], hover_fg=colors["fg"],
            command=self.toggle_pin, padx=4, pady=2,
        )
        self.btn_pin.pack(side=tk.RIGHT, padx=(0, 6))

        # ---- КНОПКА СОЗДАНИЯ ----
        self.btn_create = HoverButton(
            self.left_panel, text="+ Создать аккаунт",
            bg=colors["accent"], fg=colors["accent_fg"],
            hover_bg=colors["accent_hover"], hover_fg=colors["accent_fg"],
            font=FONT_BOLD, command=self.start_create_account, pady=8,
        )
        self.btn_create.grid(row=row, column=0, sticky="ew",
                             padx=self.PAD_X, pady=(4, 8))
        row += 1

        # ---- СЕКЦИЯ: Аккаунты ----
        self.section_accounts = SectionLabel(
            self.left_panel, text="АККАУНТЫ", font=FONT_SECTION,
            bg=colors["panel_bg"], fg=colors["muted"],
            line_color=colors["separator"]
        )
        self.section_accounts.grid(row=row, column=0, sticky="ew",
                                   padx=self.PAD_X, pady=(0, 4))
        row += 1

        # Кнопки файлов (строка)
        self.file_btn_frame = tk.Frame(self.left_panel, bg=colors["panel_bg"])
        self.file_btn_frame.grid(row=row, column=0, sticky="ew",
                                 padx=self.PAD_X, pady=(0, 4))
        row += 1

        self.btn_reload = HoverButton(
            self.file_btn_frame, text="Обновить", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.load_accounts_from_file,
        )
        self.btn_reload.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))

        self.btn_open_file = HoverButton(
            self.file_btn_frame, text="Файл", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.open_accounts_file,
        )
        self.btn_open_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.btn_open_excel = HoverButton(
            self.file_btn_frame, text="Excel", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.open_excel_file,
        )
        self.btn_open_excel.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

        self.btn_check_ban = HoverButton(
            self.file_btn_frame, text="Бан", font=FONT_SMALL,
            bg=colors["danger"], fg="white",
            hover_bg=colors["danger_hover"], hover_fg="white",
            command=self.start_ban_check,
        )
        self.btn_check_ban.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(2, 0))

        # ---- СПИСОК АККАУНТОВ (с scrollbar) ----
        self.acc_frame = tk.Frame(
            self.left_panel,
            bg=colors["surface"],
            bd=0,
            highlightthickness=1,
            highlightbackground=colors["border"],
            highlightcolor=colors["accent"],
        )
        self.acc_frame.grid(row=row, column=0, sticky="nsew",
                            padx=self.PAD_X, pady=(0, 4))
        self.left_panel.grid_rowconfigure(row, weight=1)
        row += 1

        self.acc_scrollbar = ttk.Scrollbar(
            self.acc_frame, orient=tk.VERTICAL,
            style="Dark.Vertical.TScrollbar",
        )
        self.acc_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.acc_listbox = tk.Listbox(
            self.acc_frame, height=12, exportselection=False,
            font=FONT_SMALL, activestyle="none",
            yscrollcommand=self.acc_scrollbar.set,
            relief=tk.FLAT, borderwidth=0, highlightthickness=1,
            highlightcolor=colors["accent"], highlightbackground=colors["border"],
            selectborderwidth=0,
            bg=colors["list_bg"],
            fg=colors["list_fg"],
        )
        self.acc_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.acc_scrollbar.config(command=self.acc_listbox.yview)
        self.acc_listbox.bind("<<ListboxSelect>>", self.on_account_select)

        # Контекстное меню
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(
            label="Статус: Не зарегистрирован",
            command=lambda: self.set_account_status("not_registered"),
        )
        self.context_menu.add_command(
            label="Статус: Зарегистрирован",
            command=lambda: self.set_account_status("registered"),
        )
        self.context_menu.add_command(
            label="Статус: Plus",
            command=lambda: self.set_account_status("plus"),
        )
        self.context_menu.add_command(
            label="Статус: Banned",
            command=lambda: self.set_account_status("banned"),
        )
        self.context_menu.add_command(
            label="Статус: Неверный пароль",
            command=lambda: self.set_account_status("invalid_password"),
        )
        self.acc_listbox.bind("<Button-3>", self.show_context_menu)

        # ---- СЕКЦИЯ: Действия ----
        self.section_actions = SectionLabel(
            self.left_panel, text="ДЕЙСТВИЯ", font=FONT_SECTION,
            bg=colors["panel_bg"], fg=colors["muted"],
            line_color=colors["separator"]
        )
        self.section_actions.grid(row=row, column=0, sticky="ew",
                                  padx=self.PAD_X, pady=(4, 4))
        row += 1

        # Кнопки копирования
        self.btn_frame = tk.Frame(self.left_panel, bg=colors["panel_bg"])
        self.btn_frame.grid(row=row, column=0, sticky="ew",
                            padx=self.PAD_X, pady=(0, 2))
        row += 1

        self.btn_copy_email = HoverButton(
            self.btn_frame, text="Email", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.copy_email,
        )
        self.btn_copy_email.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 2))

        self.btn_copy_pass_openai = HoverButton(
            self.btn_frame, text="OpenAI", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.copy_pass_openai,
        )
        self.btn_copy_pass_openai.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)

        self.btn_copy_pass = HoverButton(
            self.btn_frame, text="Почта", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.copy_pass,
        )
        self.btn_copy_pass.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(2, 0))

        # Кнопки инструментов
        self.tools_frame = tk.Frame(self.left_panel, bg=colors["panel_bg"])
        self.tools_frame.grid(row=row, column=0, sticky="ew",
                              padx=self.PAD_X, pady=(2, 4))
        row += 1

        self.btn_sk = HoverButton(
            self.tools_frame, text="SK", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self._show_sk_window,
        )
        self.btn_sk.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(0, 2))

        self.btn_in = HoverButton(
            self.tools_frame, text="IN", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self._show_in_window,
        )
        self.btn_in.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)

        self.btn_minesweeper = HoverButton(
            self.tools_frame, text="Сапёр", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self._show_minesweeper,
        )
        self.btn_minesweeper.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)

        self.btn_pool = HoverButton(
            self.tools_frame, text="Пул Карт", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self._show_pool_window,
        )
        self.btn_pool.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=2)

        self.btn_hotkey_settings = HoverButton(
            self.tools_frame, text="Настройки", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self._show_hotkey_settings,
        )
        self.btn_hotkey_settings.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=(2, 0))

        # ---- СЕКЦИЯ: Генератор данных ----
        self.section_gen = SectionLabel(
            self.left_panel, text="ГЕНЕРАТОР", font=FONT_SECTION,
            bg=colors["panel_bg"], fg=colors["muted"],
            line_color=colors["separator"]
        )
        self.section_gen.grid(row=row, column=0, sticky="ew",
                              padx=self.PAD_X, pady=(4, 4))
        row += 1

        self.person_frame = tk.Frame(self.left_panel, bg=colors["panel_bg"])
        self.person_frame.grid(row=row, column=0, sticky="ew",
                               padx=self.PAD_X, pady=(0, 8))
        row += 1

        self.random_name_var = tk.StringVar()
        self.random_birthdate_var = tk.StringVar()

        # Name row
        name_row = tk.Frame(self.person_frame, bg=colors["panel_bg"])
        name_row.pack(fill=tk.X, pady=2)
        tk.Label(
            name_row, text="Name", font=FONT_SMALL,
            bg=colors["panel_bg"], fg=colors["muted"], width=6, anchor="w"
        ).pack(side=tk.LEFT)
        self.entry_random_name = tk.Entry(
            name_row, textvariable=self.random_name_var, font=FONT_SMALL,
            state="readonly", width=16, relief=tk.FLAT, bd=0,
            highlightthickness=1,
        )
        self.entry_random_name.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
        self.btn_copy_random_name = HoverButton(
            name_row, text="Копировать", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.copy_random_name, padx=6, pady=2,
        )
        self.btn_copy_random_name.pack(side=tk.LEFT, padx=(2, 0))

        # Birthday row
        bdate_row = tk.Frame(self.person_frame, bg=colors["panel_bg"])
        bdate_row.pack(fill=tk.X, pady=2)
        tk.Label(
            bdate_row, text="Дата", font=FONT_SMALL,
            bg=colors["panel_bg"], fg=colors["muted"], width=6, anchor="w"
        ).pack(side=tk.LEFT)
        self.entry_random_bdate = tk.Entry(
            bdate_row, textvariable=self.random_birthdate_var, font=FONT_SMALL,
            state="readonly", width=16, relief=tk.FLAT, bd=0,
            highlightthickness=1,
        )
        self.entry_random_bdate.pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
        self.btn_copy_random_bdate = HoverButton(
            bdate_row, text="Копировать", font=FONT_SMALL,
            bg=colors["btn_bg"], fg=colors["btn_fg"],
            hover_bg=colors["btn_hover"],
            command=self.copy_random_birthdate, padx=6, pady=2,
        )
        self.btn_copy_random_bdate.pack(side=tk.LEFT, padx=(2, 0))

        # Кнопка генерации
        self.btn_generate_person = HoverButton(
            self.person_frame, text="Новые данные", font=FONT_SMALL,
            bg=colors["accent"], fg=colors["accent_fg"],
            hover_bg=colors["accent_hover"], hover_fg=colors["accent_fg"],
            command=self.generate_random_person, pady=4,
        )
        self.btn_generate_person.pack(fill=tk.X, pady=(6, 0))

    def _build_right_panel(self, colors):
        """Построение правой панели (письма)."""
        self.right_panel = tk.Frame(self.paned, bg=colors["bg"])
        self.paned.add(self.right_panel, minsize=450)

        # ---- HEADER ----
        self.header_frame = tk.Frame(self.right_panel, bg=colors["header_bg"])
        self.header_frame.pack(fill=tk.X)

        # Аккаунт-лейбл
        self.lbl_current_email = tk.Label(
            self.header_frame,
            text="Выберите аккаунт слева",
            font=FONT_TITLE,
            bg=colors["header_bg"],
            fg=colors["fg"],
            pady=12,
        )
        self.lbl_current_email.pack(side=tk.LEFT, padx=self.PAD_X, fill=tk.X, expand=True)

        # Кнопки статуса
        self.status_frame = tk.Frame(self.header_frame, bg=colors["header_bg"])
        self.status_frame.pack(side=tk.RIGHT, padx=(0, 4))

        self.btn_nr = HoverButton(
            self.status_frame, text="Не рег", font=FONT_SMALL,
            bg=STATUS_COLORS["not_registered"]["light"],
            fg=colors["btn_fg"],
            hover_bg="#e2e8f0",
            command=lambda: self.set_account_status("not_registered"),
            padx=6, pady=3,
        )
        self.btn_nr.pack(side=tk.LEFT, padx=1)

        self.btn_reg = HoverButton(
            self.status_frame, text="Рег", font=FONT_SMALL,
            bg=STATUS_COLORS["registered"]["light"],
            fg=colors["btn_fg"],
            hover_bg="#bfdbfe",
            command=lambda: self.set_account_status("registered"),
            padx=6, pady=3,
        )
        self.btn_reg.pack(side=tk.LEFT, padx=1)

        self.btn_plus = HoverButton(
            self.status_frame, text="Plus", font=FONT_SMALL,
            bg=STATUS_COLORS["plus"]["light"],
            fg=colors["btn_fg"],
            hover_bg="#9ae6b4",
            command=lambda: self.set_account_status("plus"),
            padx=6, pady=3,
        )
        self.btn_plus.pack(side=tk.LEFT, padx=1)

        # Кнопка обновления
        self.btn_refresh = HoverButton(
            self.header_frame, text="Обновить", font=FONT_SMALL,
            bg=colors["accent"], fg=colors["accent_fg"],
            hover_bg=colors["accent_hover"], hover_fg=colors["accent_fg"],
            command=self.on_manual_refresh,
            padx=12, pady=4,
        )
        self.btn_refresh.pack(side=tk.RIGHT, padx=(4, self.PAD_X))

        # Разделитель под header
        self.header_separator = tk.Frame(
            self.right_panel, bg=colors["separator"], height=1
        )
        self.header_separator.pack(fill=tk.X)

        # ---- СПИСОК ПИСЕМ ----
        self.tree_frame = tk.Frame(
            self.right_panel,
            bg=colors["surface"],
            bd=0,
            highlightthickness=1,
            highlightbackground=colors["border"],
            highlightcolor=colors["accent"],
        )
        self.tree_frame.pack(fill=tk.X, padx=self.PAD_X, pady=(8, 0))

        columns = ("sender", "subject", "date", "msg_id")
        self.tree = ttk.Treeview(
            self.tree_frame,
            columns=columns,
            displaycolumns=("sender", "subject", "date"),
            show="headings",
            height=8,
            style="Mail.Treeview",
        )
        self.tree.heading("sender", text="От кого")
        self.tree.heading("subject", text="Тема")
        self.tree.heading("date", text="Время")
        self.tree.column("sender", width=150, minwidth=100)
        self.tree.column("subject", width=300, minwidth=150)
        self.tree.column("date", width=80, anchor="center", minwidth=70)
        self.tree.column("msg_id", width=0, stretch=False)

        self.tree_scrollbar = ttk.Scrollbar(
            self.tree_frame, orient=tk.VERTICAL, command=self.tree.yview,
            style="Dark.Vertical.TScrollbar",
        )
        self.tree.configure(yscrollcommand=self.tree_scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.bind("<<TreeviewSelect>>", self.on_message_select)

        # ---- СОДЕРЖАНИЕ ПИСЬМА ----
        self.msg_header = tk.Frame(self.right_panel, bg=colors["bg"])
        self.msg_header.pack(fill=tk.X, padx=self.PAD_X, pady=(8, 0))

        self.lbl_msg_title = tk.Label(
            self.msg_header, text="Содержание письма", anchor="w",
            font=FONT_BOLD, bg=colors["bg"], fg=colors["fg"],
        )
        self.lbl_msg_title.pack(side=tk.LEFT)

        # Кнопка копирования кода (скрыта по умолчанию)
        self.btn_copy_code = HoverButton(
            self.right_panel, text="Копировать код",
            bg=colors["warning"], fg="#1a1a2e",
            hover_bg="#f6ad55", hover_fg="#1a1a2e",
            font=FONT_BOLD,
        )
        self.btn_copy_code.pack(fill=tk.X, padx=self.PAD_X, pady=4)
        self.btn_copy_code.pack_forget()

        # Текст письма
        self.msg_text_frame = tk.Frame(
            self.right_panel,
            bg=colors["surface"],
            bd=0,
            highlightthickness=1,
            highlightbackground=colors["border"],
            highlightcolor=colors["accent"],
        )
        self.msg_text_frame.pack(fill=tk.BOTH, expand=True, padx=self.PAD_X,
                                 pady=(4, self.PAD_X))

        self.msg_scrollbar = ttk.Scrollbar(
            self.msg_text_frame, orient=tk.VERTICAL,
            style="Dark.Vertical.TScrollbar",
        )
        self.msg_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.msg_text = tk.Text(
            self.msg_text_frame, wrap=tk.WORD, height=10, font=FONT_BASE,
            relief=tk.FLAT, borderwidth=0,
            highlightthickness=1,
            highlightbackground=colors["border"],
            highlightcolor=colors["accent"],
            bg=colors["text_bg"],
            fg=colors["text_fg"],
            insertbackground=colors["fg"],
            yscrollcommand=self.msg_scrollbar.set,
            padx=14,
            pady=12,
            spacing1=2,
            spacing3=2,
        )
        self.msg_text.pack(fill=tk.BOTH, expand=True)
        self.msg_scrollbar.config(command=self.msg_text.yview)
        self.msg_text.insert(tk.END, "Выберите письмо, чтобы увидеть содержимое.")

    # ================================================================
    #  WINDOW ANIMATION
    # ================================================================

    def _set_window_alpha(self, alpha):
        """Установить прозрачность окна. Возвращает True если поддерживается."""
        try:
            self.root.attributes("-alpha", alpha)
            return True
        except Exception:
            return False

    def _animate_window_open(self, step=0):
        """Плавное появление окна (fade-in)."""
        if not self._window_alpha_supported:
            return
        total = self.WINDOW_FADE_STEPS
        if step <= total:
            alpha = step / total
            self._set_window_alpha(alpha)
            self.root.after(
                self.WINDOW_FADE_DELAY,
                lambda: self._animate_window_open(step + 1),
            )

    # ================================================================
    #  NETWORKING
    # ================================================================

    def _create_http_session(self):
        """Создание HTTP сессии с настройками переподключения для устойчивости к смене VPN."""
        session = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=0.5,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["HEAD", "GET", "POST", "OPTIONS"],
            raise_on_status=False,
        )
        adapter = HTTPAdapter(
            max_retries=retry_strategy,
            pool_connections=1,
            pool_maxsize=1,
            pool_block=False,
        )
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        return session

    def _reset_http_session(self):
        """Сбросить HTTP сессию (полезно при смене VPN)."""
        try:
            if self.http_session:
                self.http_session.close()
        except Exception:
            pass
        self.http_session = self._create_http_session()

    def _make_request(self, method, url, retry_auth=True, **kwargs):
        """Выполнить HTTP запрос с обработкой ошибок сети и переподключением."""
        try:
            response = getattr(self.http_session, method)(url, timeout=10, **kwargs)
            return response
        except (
            requests.exceptions.ConnectionError,
            requests.exceptions.Timeout,
            requests.exceptions.ChunkedEncodingError,
            OSError,
        ) as e:
            print(f"[Network] Connection error: {e}")
            self._reset_http_session()
            try:
                response = getattr(self.http_session, method)(url, timeout=10, **kwargs)
                return response
            except Exception as e2:
                print(f"[Network] Retry failed: {e2}")
                if retry_auth and self.current_email and self.current_password:
                    print("[Network] Attempting re-authentication...")
                    self._try_reauth()
                return None
        except Exception as e:
            print(f"[Network] Unexpected error: {e}")
            return None

    def _try_reauth(self):
        """Попытка переавторизации при потере соединения."""
        if not self.current_email or not self.current_password:
            return False
        try:
            self.current_token = None
            if self.imap_client:
                try:
                    self.imap_client.logout()
                except Exception:
                    pass
                self.imap_client = None
            self.root.after(
                0, lambda: self.update_status("Переподключение после смены сети...")
            )
            threading.Thread(
                target=self.login_thread,
                args=(self.current_email, self.current_password),
                daemon=True,
            ).start()
            return True
        except Exception as e:
            print(f"[Reauth] Failed: {e}")
            return False

    # ================================================================
    #  HOTKEYS
    # ================================================================

    def _setup_hotkeys(self):
        """Настройка глобальных горячих клавиш."""
        self.hotkey_settings = HotkeySettings.get_instance()
        self.hotkey_settings.set_callback("email", self.copy_email)
        self.hotkey_settings.set_callback("password", self.copy_pass)
        self.hotkey_settings.set_callback(
            "paste_account", self.paste_accounts_from_clipboard
        )
        self.hotkey_settings.set_callback("copy_account", self.copy_full_account)
        self.hotkey_settings.set_callback("random_name", self.copy_random_name)
        self.hotkey_settings.set_callback("random_bdate", self.copy_random_birthdate)
        self.hotkey_settings.set_callback("refresh", self.on_manual_refresh)
        self.hotkey_settings.register_all()

    def _show_pool_window(self):
        show_pre_generator_window(self.root, theme_name=self.params.get("theme", self.DEFAULT_THEME))
        self.hotkey_settings.set_callback(
            "random_birthdate", self.copy_random_birthdate
        )
        self.hotkey_settings.register_all()

    # ================================================================
    #  CLIPBOARD OPERATIONS
    # ================================================================

    def _parse_account_line(self, raw_line):
        """Разобрать строку аккаунта из буфера или файла."""
        line = raw_line.strip()
        if not line:
            return None

        email = ""
        password_openai = ""
        password_mail = ""
        status = "not_registered"

        if " / " in line:
            parts = [part.strip() for part in line.split(" / ", 2)]
            if len(parts) >= 2:
                email = parts[0]
                passwords = parts[1]
                if ";" in passwords:
                    pwd_parts = passwords.split(";", 1)
                    password_openai = pwd_parts[0].strip()
                    password_mail = pwd_parts[1].strip()
                else:
                    password_openai = passwords.strip()
                    password_mail = passwords.strip()
                if len(parts) >= 3 and parts[2]:
                    status = parts[2]
        elif ":" in line:
            parts = line.split(":", 1)
            if len(parts) == 2:
                email = parts[0].strip()
                passwords = parts[1].strip()
                if ";" in passwords:
                    pwd_parts = passwords.split(";", 1)
                    password_openai = pwd_parts[0].strip()
                    password_mail = pwd_parts[1].strip()
                else:
                    password_openai = passwords
                    password_mail = passwords
        elif "\t" in line:
            parts = line.split("\t", 1)
            if len(parts) == 2:
                email = parts[0].strip()
                passwords = parts[1].strip()
                if ";" in passwords:
                    pwd_parts = passwords.split(";", 1)
                    password_openai = pwd_parts[0].strip()
                    password_mail = pwd_parts[1].strip()
                else:
                    password_openai = passwords
                    password_mail = passwords

        if "@" not in email:
            return None
        if not password_openai and not password_mail:
            return None
        if not password_openai:
            password_openai = password_mail
        if not password_mail:
            password_mail = password_openai

        return {
            "email": email,
            "password_openai": password_openai,
            "password_mail": password_mail,
            "password": password_mail,
            "status": status,
        }

    def paste_accounts_from_clipboard(self):
        """Вставить аккаунты из буфера обмена."""
        try:
            clipboard_text = pyperclip.paste()
            if not clipboard_text:
                self.update_status("Буфер обмена пуст")
                return

            lines = clipboard_text.strip().split("\n")
            added_count = 0

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                account = self._parse_account_line(line)
                if account:
                    email = account["email"]
                    exists = any(acc["email"] == email for acc in self.accounts_data)
                    if not exists:
                        self.accounts_data.append(account)
                        self.acc_listbox.insert(tk.END, email)
                        added_count += 1

            if added_count > 0:
                self.update_listbox_colors()
                self.save_accounts_to_file()
                self.update_status(f"Добавлено аккаунтов: {added_count}")
            else:
                self.update_status("Не найдено валидных аккаунтов для добавления")

        except Exception as e:
            self.update_status(f"Ошибка вставки: {e}")

    def copy_full_account(self):
        """Копировать полный аккаунт (email:password_openai;password_mail)."""
        selection = self.acc_listbox.curselection()
        if not selection:
            self.update_status("Выберите аккаунт для копирования")
            return

        idx = selection[0]
        if idx < len(self.accounts_data):
            acc = self.accounts_data[idx]
            password_openai = acc.get("password_openai", acc.get("password", ""))
            password_mail = acc.get("password_mail", acc.get("password", ""))

            if password_openai != password_mail:
                full_text = f"{acc['email']}:{password_openai};{password_mail}"
            else:
                full_text = f"{acc['email']}:{password_openai}"

            pyperclip.copy(full_text)
            self.update_status(f"Скопировано: {acc['email']}:***")

    def copy_email(self):
        """Копировать email выбранного аккаунта."""
        selection = self.acc_listbox.curselection()
        if not selection:
            self.update_status("Выберите аккаунт для копирования")
            return
        idx = selection[0]
        if idx < len(self.accounts_data):
            acc = self.accounts_data[idx]
            pyperclip.copy(acc["email"])
            self.update_status(f"Скопирован email: {acc['email']}")

    def copy_pass_openai(self):
        """Копировать пароль OpenAI выбранного аккаунта."""
        selection = self.acc_listbox.curselection()
        if not selection:
            self.update_status("Выберите аккаунт для копирования")
            return
        idx = selection[0]
        if idx < len(self.accounts_data):
            acc = self.accounts_data[idx]
            password = acc.get("password_openai", acc.get("password", ""))
            pyperclip.copy(password)
            self.update_status(f"Скопирован пароль OpenAI для: {acc['email']}")

    def copy_pass(self):
        """Копировать пароль от почты выбранного аккаунта."""
        selection = self.acc_listbox.curselection()
        if not selection:
            self.update_status("Выберите аккаунт для копирования")
            return
        idx = selection[0]
        if idx < len(self.accounts_data):
            acc = self.accounts_data[idx]
            password = acc.get("password_mail", acc.get("password", ""))
            pyperclip.copy(password)
            self.update_status(f"Скопирован пароль почты для: {acc['email']}")

    def copy_random_name(self):
        """Копировать случайное имя."""
        name = self.random_name_var.get()
        if name:
            pyperclip.copy(name)
            self.update_status(f"Скопировано имя: {name[:20]}")

    def copy_random_birthdate(self):
        """Копировать случайную дату рождения."""
        bdate = self.random_birthdate_var.get()
        if bdate:
            pyperclip.copy(bdate)
            self.update_status(f"Скопирована дата: {bdate}")

    def copy_code_to_clipboard(self, code):
        """Копирование кода в буфер."""
        pyperclip.copy(code)
        self.status_var.set(f"Код {code} скопирован!")

    # ================================================================
    #  WINDOW HELPERS
    # ================================================================

    def toggle_pin(self):
        """Переключение режима 'Поверх всех окон'."""
        self.is_pinned = not self.is_pinned
        self.root.wm_attributes("-topmost", self.is_pinned)
        colors = THEMES[self.params.get("theme", "light")]
        if self.is_pinned:
            self.btn_pin.update_colors(
                bg=colors["accent"], fg=colors["accent_fg"],
                hover_bg=colors["accent_hover"], hover_fg=colors["accent_fg"]
            )
        else:
            self.btn_pin.update_colors(
                bg=colors["panel_bg"], fg=colors["muted"],
                hover_bg=colors["btn_hover"], hover_fg=colors["fg"]
            )

    def _show_hotkey_settings(self):
        """Открыть окно настроек горячих клавиш."""
        def on_save(new_hotkeys):
            self.hotkey_settings.register_all()
        theme_name = self.params.get("theme", "light")
        show_settings_window(self.root, theme_name, on_save=on_save)

    def _show_sk_window(self):
        theme_name = self.params.get("theme", "light")
        show_sk_window(self.root, theme_name)

    def _show_in_window(self):
        theme_name = self.params.get("theme", "light")
        show_in_window(self.root, theme_name)

    def _show_minesweeper(self):
        theme_name = self.params.get("theme", "light")
        show_minesweeper(self.root, theme_name)

    # ================================================================
    #  RANDOM DATA GENERATION
    # ================================================================

    def generate_random_person(self):
        """Генерация случайных данных о человеке."""
        name = self.fake.first_name()
        self.random_name_var.set(name)

        start_date = datetime(1975, 1, 1)
        end_date = datetime(2004, 12, 31)
        days_between = (end_date - start_date).days
        random_days = random.randint(0, days_between)
        birthdate = start_date + timedelta(days=random_days)
        self.random_birthdate_var.set(birthdate.strftime("%d.%m.%Y"))
        self.update_status(f"Сгенерировано: {name}")

    # ================================================================
    #  SOUND
    # ================================================================

    def play_notification_sound(self, count=1):
        """Проигрывает звук при появлении новых писем."""
        def _beep():
            for _ in range(max(1, count)):
                try:
                    if platform.system() == "Windows":
                        winsound.MessageBeep(winsound.MB_ICONASTERISK)
                    else:
                        print("\a", end="", flush=True)
                except Exception:
                    pass
                time.sleep(0.1)
        threading.Thread(target=_beep, daemon=True).start()

    # ================================================================
    #  STATUS
    # ================================================================

    def update_status(self, text):
        """Безопасное обновление статуса из другого потока."""
        self.root.after(0, lambda: self.status_var.set(text))

    # ================================================================
    #  FILE OPERATIONS
    # ================================================================

    def open_accounts_file(self):
        """Открывает файл аккаунтов в системном редакторе."""
        try:
            if not os.path.exists(ACCOUNTS_FILE):
                with open(ACCOUNTS_FILE, "w") as f:
                    pass
            os.startfile(ACCOUNTS_FILE)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть файл:\n{e}")

    def open_excel_file(self):
        """Открывает Excel файл аккаунтов."""
        try:
            if not os.path.exists(EXCEL_FILE):
                self.save_accounts_to_excel()
            os.startfile(EXCEL_FILE)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть Excel файл:\n{e}")

    # ================================================================
    #  BAN CHECK
    # ================================================================

    def start_ban_check(self):
        """Запуск проверки всех аккаунтов на бан OpenAI."""
        if not self.accounts_data:
            messagebox.showwarning("Внимание", "Нет аккаунтов для проверки")
            return

        total = len(self.accounts_data)
        if not messagebox.askyesno(
            "Проверка бана",
            f"Проверить {total} аккаунтов на бан OpenAI?\n\n"
            "Используется многопоточность для ускорения.\n"
            "Аккаунты с письмом 'Access Deactivated' будут помечены как забаненные.",
        ):
            return

        self.btn_check_ban.config(state=tk.DISABLED)
        self.btn_check_ban.config(text="Проверка...")

        # Больше потоков — операции I/O-bound (сеть), CPU не загружен
        recommended_threads = min(60, max(8, total // 3))
        self.ban_check_threads = max(1, min(total, recommended_threads))
        self.ban_check_lock = threading.Lock()

        # Сброс thread-local хранилища от предыдущего запуска
        self._ban_thread_local = threading.local()

        self._ban_start_time = time.time()
        self._create_progress_window(total)
        threading.Thread(target=self.ban_check_thread, daemon=True).start()

    def _create_progress_window(self, total):
        """Создание окна с прогресс баром."""
        self.progress_window = tk.Toplevel(self.root)
        self.progress_window.title("Проверка на бан OpenAI")
        self.progress_window.geometry("480x200")
        self.progress_window.resizable(False, False)
        self.progress_window.transient(self.root)
        self.progress_window.grab_set()

        self.progress_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - 480) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 200) // 2
        self.progress_window.geometry(f"+{x}+{y}")

        theme = self.params.get("theme", "light")
        colors = THEMES[theme]
        self.progress_window.config(bg=colors["panel_bg"])

        self.progress_title = tk.Label(
            self.progress_window,
            text="Проверка аккаунтов на бан...",
            font=FONT_BOLD,
            bg=colors["panel_bg"],
            fg=colors["fg"],
        )
        self.progress_title.pack(pady=(20, 8))

        self.progress_label = tk.Label(
            self.progress_window,
            text=f"Подготовка... 0/{total}",
            font=FONT_SMALL,
            bg=colors["panel_bg"],
            fg=colors["muted"],
        )
        self.progress_label.pack(pady=4)

        style = ttk.Style()
        style.configure(
            "ban.Horizontal.TProgressbar",
            troughcolor=colors.get("entry_bg", "#e5e7eb"),
            background=colors["danger"],
        )
        self.progress_bar = ttk.Progressbar(
            self.progress_window,
            orient="horizontal",
            length=420,
            mode="determinate",
            maximum=total,
            style="ban.Horizontal.TProgressbar",
        )
        self.progress_bar.pack(pady=10)

        self.progress_stats = tk.Label(
            self.progress_window,
            text="Забанено: 0 | Проверено: 0",
            font=FONT_SMALL,
            bg=colors["panel_bg"],
            fg=colors["fg"],
        )
        self.progress_stats.pack(pady=4)

        self.ban_check_cancelled = False
        self.btn_cancel_ban = HoverButton(
            self.progress_window,
            text="Отмена",
            font=FONT_SMALL,
            bg=colors["btn_bg"],
            fg=colors["btn_fg"],
            hover_bg=colors["danger"],
            hover_fg="white",
            command=self._cancel_ban_check,
        )
        self.btn_cancel_ban.pack(pady=8)

        self.progress_window.protocol("WM_DELETE_WINDOW", self._cancel_ban_check)

    def _cancel_ban_check(self):
        """Отмена проверки бана."""
        self.ban_check_cancelled = True
        self.progress_label.config(text="Отмена...")

    def _update_progress(self, current, total, email, banned_count, checked_count):
        """Обновление прогресс бара."""
        try:
            if hasattr(self, "progress_window") and self.progress_window.winfo_exists():
                self.progress_bar["value"] = current
                elapsed = time.time() - getattr(self, "_ban_start_time", time.time())
                speed = checked_count / max(elapsed, 0.1)
                remaining = (total - checked_count) / max(speed, 0.1)
                self.progress_label.config(
                    text=f"{email[:30]}... ({current}/{total})"
                )
                self.progress_stats.config(
                    text=f"Забанено: {banned_count} | Проверено: {checked_count} | "
                         f"{speed:.1f} акк/с | ~{remaining:.0f}с"
                )
        except tk.TclError:
            pass  # окно уже закрыто

    def _get_ban_thread_session(self):
        """HTTP сессия для текущего потока проверки (быстрая, минимальные retry)."""
        session = getattr(self._ban_thread_local, "session", None)
        if session is not None:
            return session

        session = requests.Session()
        # Минимальные retry — не тратим время на повторы при бан-чеке
        retry_strategy = Retry(
            total=1,
            backoff_factor=0.1,
            status_forcelist=[502, 503, 504],
            allowed_methods=["HEAD", "GET", "POST"],
            raise_on_status=False,
        )
        adapter = HTTPAdapter(
            max_retries=retry_strategy,
            pool_connections=2,
            pool_maxsize=2,
            pool_block=False,
        )
        session.mount("http://", adapter)
        session.mount("https://", adapter)

        self._ban_thread_local.session = session
        with self._ban_thread_sessions_lock:
            self._ban_thread_sessions.append(session)
        return session

    def _close_ban_thread_sessions(self):
        """Закрывает все HTTP сессии, созданные потоками ban-check."""
        with self._ban_thread_sessions_lock:
            sessions = self._ban_thread_sessions[:]
            self._ban_thread_sessions = []
        for session in sessions:
            try:
                session.close()
            except Exception:
                pass

    def _get_ban_imap_hosts(self, domain):
        """Возвращает IMAP хосты в порядке приоритета."""
        with self._ban_imap_host_lock:
            cached = self._ban_imap_host_cache.get(domain)
        candidates = [cached, "imap.firstmail.ltd", f"imap.{domain}"]
        hosts = []
        for host in candidates:
            if host and host not in hosts:
                hosts.append(host)
        return hosts

    def _remember_ban_imap_host(self, domain, host):
        """Запоминает рабочий IMAP host для домена."""
        with self._ban_imap_host_lock:
            self._ban_imap_host_cache[domain] = host

    def ban_check_thread(self):
        """Поток проверки всех аккаунтов на бан с многопоточностью."""
        banned_count = 0
        invalid_pass_count = 0
        error_count = 0
        checked_count = 0
        total = len(self.accounts_data)
        start_time = time.time()
        last_ui_update = [0.0]  # список для замыкания; пишется только из main loop

        def check_single_account(idx, account):
            email_addr = account.get("email", "")
            password = account.get("password_mail", account.get("password", ""))
            old_status = account.get("status", "not_registered")

            if not email_addr or not password:
                return (idx, email_addr, None, None, True)
            if old_status in ("banned", "invalid_password"):
                return (idx, email_addr, None, None, True)

            try:
                result, reason = self._check_account_for_ban_threadsafe(email_addr, password)
                return (idx, email_addr, result, reason, False)
            except Exception as e:
                return (idx, email_addr, "error", str(e), False)

        try:
            with ThreadPoolExecutor(max_workers=self.ban_check_threads) as executor:
                futures = {
                    executor.submit(check_single_account, idx, account): idx
                    for idx, account in enumerate(self.accounts_data)
                }

                for future in as_completed(futures):
                    if self.ban_check_cancelled:
                        executor.shutdown(wait=False, cancel_futures=True)
                        break

                    try:
                        idx, email_addr, result, reason, skipped = future.result()

                        with self.ban_check_lock:
                            if skipped or not email_addr:
                                checked_count += 1
                                continue

                            if result == "banned":
                                self.accounts_data[idx]["status"] = "banned"
                                banned_count += 1
                            elif result == "invalid_password":
                                self.accounts_data[idx]["status"] = "invalid_password"
                                invalid_pass_count += 1
                            elif result == "error":
                                error_count += 1

                            checked_count += 1

                            # Снимаем снимок счётчиков под локом
                            _checked = checked_count
                            _banned = banned_count
                            _email = email_addr

                        # Обновляем UI не чаще чем раз в 150мс
                        now = time.monotonic()
                        if now - last_ui_update[0] > 0.15:
                            last_ui_update[0] = now
                            self.root.after(
                                0,
                                lambda c=_checked,
                                e=_email,
                                b=_banned: self._update_progress(c, total, e, b, c),
                            )

                    except Exception as e:
                        print(f"[BAN] Future error: {e}")
                        with self.ban_check_lock:
                            checked_count += 1
        finally:
            self._close_ban_thread_sessions()

        elapsed_time = time.time() - start_time
        speed = total / max(elapsed_time, 0.1)
        print(f"[BAN] Завершено за {elapsed_time:.1f}с ({speed:.1f} акк/с), ошибок: {error_count}")

        self.root.after(
            0,
            lambda: self._on_ban_check_complete(
                checked_count, banned_count, invalid_pass_count, error_count
            ),
        )

    @staticmethod
    def _is_openai_ban_message(sender, subject):
        """Проверяет, является ли письмо уведомлением о бане OpenAI."""
        sender = sender.lower()
        subject = subject.lower()
        if "openai" not in sender:
            return False
        ban_keywords = [
            "access deactivated",
            "deactivated",
            "account suspended",
            "account disabled",
            "account has been disabled",
            "account has been deactivated",
            "suspended",
            "violation",
        ]
        return any(kw in subject for kw in ban_keywords)

    @staticmethod
    def _extract_sender_address(from_field):
        """Извлечение email из поля From (поддерживает 'Name <email>' и dict)."""
        if isinstance(from_field, dict):
            return from_field.get("address", "")
        # IMAP возвращает строку вида "OpenAI <noreply@tm.openai.com>"
        s = str(from_field)
        if "<" in s and ">" in s:
            return s[s.index("<") + 1:s.index(">")]
        return s

    def _check_account_for_ban_threadsafe(self, email_addr, password):
        """Потокобезопасная проверка одного аккаунта на бан OpenAI."""
        session = self._get_ban_thread_session()
        domain = email_addr.split("@")[-1]

        # Ensure domain list is loaded; mail.tm may use third-party domains
        # (e.g. dollicons.com) that don't end with "mail.tm".
        if not self.mail_tm_domains:
            self.load_mail_tm_domains()
        is_mail_tm = domain in self.mail_tm_domains or domain.endswith("mail.tm")

        if is_mail_tm:
            try:
                payload = {"address": email_addr, "password": password}
                # Короткие таймауты: 3с connect, 5с read
                res = session.post(f"{API_URL}/token", json=payload, timeout=(3, 5))

                if res.status_code == 401:
                    return ("invalid_password", "wrong_credentials")
                if res.status_code != 200:
                    return ("error", f"auth_failed_{res.status_code}")

                token = res.json().get("token")
                if not token:
                    return ("error", "no_token")

                headers = {"Authorization": f"Bearer {token}"}
                res = session.get(f"{API_URL}/messages", headers=headers, timeout=(3, 5))

                if res.status_code != 200:
                    return ("error", "messages_failed")

                messages = res.json().get("hydra:member", [])

                for msg in messages:
                    sender = self._extract_sender_address(msg.get("from", {}))
                    subject = msg.get("subject", "")

                    if self._is_openai_ban_message(sender, subject):
                        return ("banned", "access_deactivated")

                return ("ok", "no_ban_found")

            except requests.exceptions.RequestException as e:
                return ("error", str(e))

        # IMAP проверка — короткий таймаут
        imap_client = None
        any_host_reached = False
        try:
            for host in self._get_ban_imap_hosts(domain):
                try:
                    client = IMAPClient(host=host, timeout=5)
                    if client.login(email_addr, password):
                        imap_client = client
                        self._remember_ban_imap_host(domain, host)
                        break
                    # Логин не прошёл, но сервер ответил — значит пароль неверный
                    any_host_reached = True
                    client.logout()
                except (OSError, ConnectionError, TimeoutError):
                    # Хост недоступен — пробуем следующий
                    continue

            if not imap_client:
                if any_host_reached:
                    return ("invalid_password", "imap_login_failed")
                return ("error", "imap_connection_failed")

            # Проверяем только 15 последних — бан-письмо обычно среди свежих
            messages = imap_client.get_messages(limit=15)

            for msg in messages:
                sender = self._extract_sender_address(msg.get("from", {}))
                subject = msg.get("subject", "")

                if self._is_openai_ban_message(sender, subject):
                    return ("banned", "access_deactivated")

            return ("ok", "no_ban_found")

        except Exception as e:
            return ("error", str(e))
        finally:
            if imap_client:
                imap_client.logout()

    def _on_ban_check_complete(self, checked, banned, invalid_pass=0, errors=0):
        """Завершение проверки бана."""
        if hasattr(self, "progress_window") and self.progress_window.winfo_exists():
            self.progress_window.destroy()

        self.ban_check_cancelled = False

        self.btn_check_ban.config(state=tk.NORMAL, text="Бан")
        self.update_listbox_colors()
        self.save_accounts_to_file()

        msg = (
            f"Проверка завершена!\n\n"
            f"Проверено: {checked}\n"
            f"Забанено: {banned}\n"
            f"Неверный пароль: {invalid_pass}"
        )
        if errors > 0:
            msg += f"\nОшибки (сеть/таймаут): {errors}"
        if banned > 0 or invalid_pass > 0:
            messagebox.showwarning("Результаты проверки", msg)
        else:
            messagebox.showinfo("Результаты проверки", msg)

        status_msg = f"Проверка завершена. Забанено: {banned}, Неверный пароль: {invalid_pass}"
        if errors > 0:
            status_msg += f", Ошибки: {errors}"
        self.update_status(status_msg)

    # ================================================================
    #  EXCEL
    # ================================================================

    def save_accounts_to_excel(self):
        """Сохраняет данные аккаунтов в Excel файл."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Аккаунты"

            headers = ["Логин/Пароль", "Логин", "Пароль"]
            header_fill = PatternFill(
                start_color="4CAF50", end_color="4CAF50", fill_type="solid"
            )
            header_font_white = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            data_font = Font(name="Arial", size=10)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 35
            ws.column_dimensions["C"].width = 20

            status_fills = {
                "not_registered": PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                ),
                "registered": PatternFill(
                    start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                ),
                "plus": PatternFill(
                    start_color="46BDC6", end_color="46BDC6", fill_type="solid"
                ),
                "banned": PatternFill(
                    start_color="FECACA", end_color="FECACA", fill_type="solid"
                ),
                "invalid_password": PatternFill(
                    start_color="E9D5FF", end_color="E9D5FF", fill_type="solid"
                ),
            }

            for row, account in enumerate(self.accounts_data, 2):
                email = account.get("email", "")
                password = account.get("password", "")
                status = account.get("status", "not_registered")

                ws.cell(row=row, column=1, value=f"{email} / {password}")
                ws.cell(row=row, column=2, value=email)
                ws.cell(row=row, column=3, value=password)
                row_fill = status_fills.get(status, status_fills["not_registered"])
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = row_fill
                    cell.font = data_font

            wb.save(EXCEL_FILE)

        except Exception as e:
            print(f"Ошибка сохранения Excel: {e}")

    # ================================================================
    #  ACCOUNT LOADING / SAVING
    # ================================================================

    def load_accounts_from_file(self):
        """Загрузка аккаунтов из файла."""
        self.acc_listbox.delete(0, tk.END)
        self.accounts_data = []

        if os.path.exists(ACCOUNTS_FILE):
            try:
                needs_save = False

                with open(ACCOUNTS_FILE, "r", encoding="utf-8") as f:
                    lines = f.readlines()

                for line in lines:
                    account = self._parse_account_line(line)
                    if not account:
                        continue

                    if ":" in line or "\t" in line:
                        needs_save = True

                    self.accounts_data.append(account)
                    self.acc_listbox.insert(tk.END, account["email"])

                if needs_save:
                    self.save_accounts_to_file()
                    self.update_status(
                        f"Аккаунты конвертированы и загружены: {len(self.accounts_data)}"
                    )
                else:
                    self.save_accounts_to_excel()
                    self.update_status(
                        f"Загружено аккаунтов: {len(self.accounts_data)}"
                    )

                self.update_listbox_colors()

            except Exception as e:
                messagebox.showerror("Ошибка чтения файла", str(e))
        else:
            self.update_status("Файл accounts.txt не найден")

    def save_accounts_to_file(self):
        """Сохранение аккаунтов в файл."""
        try:
            with open(ACCOUNTS_FILE, "w", encoding="utf-8") as f:
                for item in self.accounts_data:
                    password_openai = item.get(
                        "password_openai", item.get("password", "")
                    )
                    password_mail = item.get("password_mail", item.get("password", ""))

                    if password_openai != password_mail:
                        passwords = f"{password_openai};{password_mail}"
                    else:
                        passwords = password_openai

                    line = f"{item['email']} / {passwords} / {item['status']}\n"
                    f.write(line)
            self.save_accounts_to_excel()
        except Exception as e:
            messagebox.showerror("Ошибка сохранения", str(e))

    # ================================================================
    #  ACCOUNT CREATION
    # ================================================================

    def start_create_account(self):
        """Запуск создания аккаунта."""
        self.btn_create.config(state=tk.DISABLED)
        self.update_status("Регистрация... (Подождите)")
        threading.Thread(target=self.create_account_thread, daemon=True).start()

    def create_account_thread(self):
        """Поток создания аккаунта."""
        try:
            domain_res = self._make_request(
                "get", f"{API_URL}/domains", retry_auth=False
            )
            if not domain_res or domain_res.status_code != 200:
                error_msg = (
                    "Сетевая ошибка"
                    if not domain_res
                    else f"Код: {domain_res.status_code}"
                )
                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "Ошибка", f"Не удалось получить список доменов\n{error_msg}"
                    ),
                )
                self.root.after(0, lambda: self.btn_create.config(state=tk.NORMAL))
                return

            domains = domain_res.json()["hydra:member"]
            domain = random.choice(domains)["domain"]

            username = "".join(
                random.choice(string.ascii_lowercase + string.digits) for _ in range(10)
            )
            chars = string.ascii_letters + string.digits
            password = "".join(random.choice(chars) for _ in range(12))

            email = f"{username}@{domain}"

            payload = {"address": email, "password": password}
            res = self._make_request(
                "post", f"{API_URL}/accounts", retry_auth=False, json=payload
            )

            if not res:
                self.update_status("Сетевая ошибка при регистрации")
                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "Ошибка", "Сетевая ошибка при регистрации"
                    ),
                )
            elif res.status_code == 201:
                self.root.after(0, lambda: self._on_account_created(email, password))
            else:
                self.update_status("Ошибка регистрации")
                self.root.after(
                    0,
                    lambda: messagebox.showerror(
                        "Ошибка", f"Код: {res.status_code}\n{res.text}"
                    ),
                )
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Ошибка", str(e)))
        finally:
            self.root.after(0, lambda: self.btn_create.config(state=tk.NORMAL))

    def _on_account_created(self, email, password):
        """Обработка созданного аккаунта."""
        self.accounts_data.append(
            {
                "email": email,
                "password_openai": password,
                "password_mail": password,
                "password": password,
                "status": "not_registered",
            }
        )

        self.acc_listbox.insert(tk.END, email)
        self.update_listbox_colors()

        self.acc_listbox.selection_clear(0, tk.END)
        self.acc_listbox.selection_set(tk.END)

        self.save_accounts_to_file()
        self.status_var.set(f"Создан: {email}")
        self.on_account_select(None)

    # ================================================================
    #  DOMAIN LOADING
    # ================================================================

    def load_mail_tm_domains(self):
        """Загрузка доменов mail.tm."""
        try:
            res = self._make_request("get", f"{API_URL}/domains", retry_auth=False)
            if res and res.status_code == 200:
                data = res.json()["hydra:member"]
                self.mail_tm_domains = [d["domain"] for d in data]
                print(f"[*] Loaded {len(self.mail_tm_domains)} mail.tm domains")
        except:
            pass

    # ================================================================
    #  THEME
    # ================================================================

    def on_theme_toggle_click(self, is_on):
        """Обработка переключения темы."""
        self.set_theme("dark" if is_on else "light")

    def set_theme(self, theme_name, animate=True):
        """Установка темы оформления."""
        self.params["theme"] = theme_name
        colors = THEMES[theme_name]
        accent_bg = colors["accent"]
        accent_fg = colors["accent_fg"]

        # Theme toggle state
        if hasattr(self, "theme_toggle"):
            self.theme_toggle.config(bg=colors["panel_bg"])
            if theme_name == "dark":
                self.theme_toggle.update_colors(
                    bg_on=colors["accent"], bg_off=colors["btn_bg"],
                    handle_color=colors["entry_fg"],
                    handle_outline=colors["border"],
                    shadow_color=colors["bg"],
                )
            else:
                self.theme_toggle.update_colors(
                    bg_on=colors["accent"], bg_off="#cbd5e0",
                    handle_color="#ffffff",
                    handle_outline="#e2e8f0",
                    shadow_color="#d4d4d4",
                )
            self.theme_toggle.set_state(theme_name == "dark")

        # Theme icon
        if hasattr(self, "lbl_theme_icon"):
            self.lbl_theme_icon.config(
                text="🌙" if theme_name == "dark" else "☀",
                bg=colors["panel_bg"], fg=colors["muted"]
            )

        # Root
        self.root.config(bg=colors["bg"])
        self.root_container.config(bg=colors["bg"])
        self.paned.config(bg=colors["separator"])
        self.status_bar.config(bg=colors["status_bg"], fg=colors["status_fg"])

        # App title
        if hasattr(self, "lbl_app_title"):
            self.lbl_app_title.config(bg=colors["panel_bg"], fg=colors["accent"])

        # Left Panel
        self.left_panel.config(bg=colors["panel_bg"])
        self.left_header.config(bg=colors["panel_bg"])
        if hasattr(self, "theme_frame"):
            self.theme_frame.config(bg=colors["panel_bg"])

        # Pin button
        if hasattr(self, "btn_pin"):
            if self.is_pinned:
                self.btn_pin.update_colors(
                    bg=colors["accent"], fg=colors["accent_fg"],
                    hover_bg=colors["accent_hover"], hover_fg=colors["accent_fg"]
                )
            else:
                self.btn_pin.update_colors(
                    bg=colors["panel_bg"], fg=colors["muted"],
                    hover_bg=colors["btn_hover"], hover_fg=colors["fg"]
                )

        # Create button
        self.btn_create.update_colors(
            bg=colors["accent"], fg=accent_fg,
            hover_bg=colors["accent_hover"], hover_fg=accent_fg,
        )

        # Section labels
        for section in [self.section_accounts, self.section_actions, self.section_gen]:
            section.update_colors(colors["panel_bg"], colors["muted"],
                                  colors["separator"])

        # File buttons frame
        self.file_btn_frame.config(bg=colors["panel_bg"])
        self.btn_frame.config(bg=colors["panel_bg"])
        self.tools_frame.config(bg=colors["panel_bg"])

        # Generic buttons (left panel)
        generic_btns = [
            self.btn_reload, self.btn_open_file, self.btn_open_excel,
            self.btn_copy_email, self.btn_copy_pass_openai, self.btn_copy_pass,
            self.btn_sk, self.btn_in, self.btn_minesweeper, self.btn_hotkey_settings,
            self.btn_copy_random_name, self.btn_copy_random_bdate,
        ]
        for btn in generic_btns:
            btn.update_colors(
                bg=colors["btn_bg"], fg=colors["btn_fg"],
                hover_bg=colors["btn_hover"],
            )

        # Ban button
        self.btn_check_ban.update_colors(
            bg=colors["danger"], fg="white",
            hover_bg=colors["danger_hover"], hover_fg="white",
        )

        # Generate button
        self.btn_generate_person.update_colors(
            bg=colors["accent"], fg=accent_fg,
            hover_bg=colors["accent_hover"], hover_fg=accent_fg,
        )

        # Person frame
        self.person_frame.config(bg=colors["panel_bg"])
        for child in self.person_frame.winfo_children():
            if isinstance(child, tk.Frame):
                child.config(bg=colors["panel_bg"])
                for subchild in child.winfo_children():
                    if isinstance(subchild, tk.Label):
                        subchild.config(bg=colors["panel_bg"], fg=colors["muted"])
                    elif isinstance(subchild, tk.Entry):
                        subchild.config(
                            readonlybackground=colors["entry_bg"],
                            fg=colors["entry_fg"],
                            highlightbackground=colors["border"],
                            highlightcolor=colors["accent"],
                        )

        # Listbox
        self.acc_listbox.config(
            bg=colors["list_bg"],
            fg=colors["list_fg"],
            selectbackground=accent_bg,
            selectforeground=accent_fg,
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=colors["border"],
            highlightcolor=colors["accent"],
        )
        self.update_listbox_colors()

        # Right Panel
        self.right_panel.config(bg=colors["bg"])
        self.header_frame.config(bg=colors["header_bg"])
        if hasattr(self, "header_separator"):
            self.header_separator.config(bg=colors["separator"])
        if hasattr(self, "tree_frame"):
            self.tree_frame.config(bg=colors["bg"])
        if hasattr(self, "msg_header"):
            self.msg_header.config(bg=colors["bg"])
        if hasattr(self, "msg_text_frame"):
            self.msg_text_frame.config(bg=colors["bg"])
        self.status_frame.config(bg=colors["header_bg"])
        self.lbl_current_email.config(bg=colors["header_bg"], fg=colors["fg"])
        self.lbl_msg_title.config(bg=colors["bg"], fg=colors["fg"])

        # Refresh button
        self.btn_refresh.update_colors(
            bg=accent_bg, fg=accent_fg,
            hover_bg=colors["accent_hover"], hover_fg=accent_fg,
        )

        # Copy code button
        self.btn_copy_code.update_colors(
            bg=colors["warning"], fg="#1a1a2e",
            hover_bg="#f6ad55", hover_fg="#1a1a2e",
        )

        # Status buttons
        status_btn_fg = colors["btn_fg"]
        self.btn_nr.update_colors(
            bg=STATUS_COLORS["not_registered"][theme_name],
            fg=status_btn_fg,
            hover_bg=colors["btn_hover"],
        )
        self.btn_reg.update_colors(
            bg=STATUS_COLORS["registered"][theme_name],
            fg=status_btn_fg,
            hover_bg="#bfdbfe" if theme_name == "light" else "#25558f",
        )
        self.btn_plus.update_colors(
            bg=STATUS_COLORS["plus"][theme_name],
            fg=status_btn_fg,
            hover_bg="#9ae6b4" if theme_name == "light" else "#236052",
        )

        # Text
        self.msg_text.config(
            bg=colors["text_bg"],
            fg=colors["text_fg"],
            insertbackground=colors["fg"],
            selectbackground=accent_bg,
            selectforeground=accent_fg,
            relief=tk.FLAT,
            borderwidth=0,
            highlightthickness=1,
            highlightbackground=colors["border"],
            highlightcolor=colors["accent"],
        )

        # Treeview Style
        style = ttk.Style()
        selected_design = (
            self.design_var.get() if hasattr(self, "design_var") else style.theme_use()
        )
        if theme_name == "dark" and "clam" in style.theme_names():
            selected_design = "clam"
        if selected_design not in style.theme_names():
            selected_design = (
                "default" if "default" in style.theme_names() else style.theme_use()
            )
        try:
            style.theme_use(selected_design)
        except Exception:
            selected_design = style.theme_use()
        if hasattr(self, "design_var"):
            self.design_var.set(selected_design)

        style.configure(
            "Mail.Treeview",
            background=colors["list_bg"],
            foreground=colors["list_fg"],
            fieldbackground=colors["list_bg"],
            rowheight=28,
            borderwidth=0,
            relief="flat",
            # Убираем белые рамки Treeview.field в clam-теме
            bordercolor=colors["list_bg"],
            lightcolor=colors["list_bg"],
            darkcolor=colors["list_bg"],
        )
        heading_border = colors["separator"] if theme_name == "dark" else ""
        style.configure(
            "Mail.Treeview.Heading",
            background=colors["header_bg"],
            foreground=colors["fg"],
            relief="flat",
            font=FONT_SMALL,
            borderwidth=1 if theme_name == "light" else 0,
            padding=(8, 4),
            # Перекрашиваем все границы heading (clam рисует их белыми)
            bordercolor=heading_border or colors["border"],
            lightcolor=colors["header_bg"],
            darkcolor=colors["header_bg"],
        )
        style.map(
            "Mail.Treeview.Heading",
            background=[("active", colors["header_bg"]), ("pressed", colors["header_bg"])],
            foreground=[("active", colors["fg"]), ("pressed", colors["fg"])],
            relief=[("active", "flat"), ("pressed", "flat")],
        )
        style.map(
            "Mail.Treeview",
            background=[
                ("selected", accent_bg),
                ("!selected", colors["list_bg"]),
            ],
            foreground=[
                ("selected", accent_fg),
                ("!selected", colors["list_fg"]),
            ],
        )

        # Убираем border из layout — оставляем только treearea (без рамки)
        style.layout("Mail.Treeview", [
            ("Treeview.treearea", {"sticky": "nswe"})
        ])

        self.tree.configure(style="Mail.Treeview")

        # Scrollbars (ttk — обновляем стиль ПОСЛЕ theme_use)
        style.configure(
            "Dark.Vertical.TScrollbar",
            background=colors["btn_bg"],
            troughcolor=colors["panel_bg"],
            bordercolor=colors["panel_bg"],
            arrowcolor=colors["fg"],
            lightcolor=colors["panel_bg"],
            darkcolor=colors["panel_bg"],
            gripcount=0,
            borderwidth=0,
            width=12,
        )
        style.map(
            "Dark.Vertical.TScrollbar",
            background=[("active", colors["btn_hover"]), ("pressed", colors["btn_hover"])],
        )

    def on_design_change(self, event=None):
        """Изменение дизайна (ttk theme)."""
        selected = self.design_var.get()
        style = ttk.Style()
        try:
            style.theme_use(selected)
            self.update_status(f"Дизайн изменен: {selected}")
        except Exception as e:
            self.update_status(f"Ошибка смены дизайна: {e}")
        self.set_theme(self.params.get("theme", "light"))

    def update_listbox_colors(self):
        """Обновление цветов списка аккаунтов."""
        theme = self.params.get("theme", "light")
        colors = THEMES[theme]
        for i in range(self.acc_listbox.size()):
            if i < len(self.accounts_data):
                status = self.accounts_data[i].get("status", "not_registered")
                color = STATUS_COLORS.get(status, {}).get(theme, colors["list_bg"])
                fg_color = colors["list_fg"]
                self.acc_listbox.itemconfig(i, {"bg": color, "fg": fg_color})

    # ================================================================
    #  ACCOUNT SELECTION / EMAIL
    # ================================================================

    def on_account_select(self, event):
        """Выбор аккаунта."""
        selection = self.acc_listbox.curselection()
        if not selection:
            return

        idx = selection[0]
        if idx >= len(self.accounts_data):
            return

        acc = self.accounts_data[idx]
        email = acc.get("email", "")
        password = acc.get("password_mail", acc.get("password", ""))

        if not email or not password:
            return

        self.lbl_current_email.config(text=email)
        self.last_message_ids = set()

        for item in self.tree.get_children():
            self.tree.delete(item)

        self.msg_text.delete(1.0, tk.END)
        self.msg_text.insert(tk.END, "Загрузка...")

        self.update_status("Авторизация...")
        threading.Thread(
            target=self.login_thread, args=(email, password), daemon=True
        ).start()

    def login_thread(self, email_addr, password):
        """Поток авторизации."""
        domain = email_addr.split("@")[-1]
        self.current_token = None
        if self.imap_client:
            try:
                self.imap_client.logout()
            except Exception:
                pass
            self.imap_client = None

        self.current_email = email_addr
        self.current_password = password
        self._reset_http_session()

        success = False

        # Always try mail.tm API first — domains may not end with "mail.tm"
        # (e.g. dollicons.com) and the async domain list may not be loaded yet.
        try:
            payload = {"address": email_addr, "password": password}
            res = self._make_request(
                "post", f"{API_URL}/token", retry_auth=False, json=payload
            )
            if res and res.status_code == 200:
                self.current_token = res.json()["token"]
                self.account_type = "api"
                success = True
            elif res:
                print(f"API Login failed: {res.status_code}")
            else:
                print("API Login failed: network error")
        except Exception as e:
            print(f"API Error: {e}")

        if not success:
            if self.imap_client:
                self.imap_client.logout()

            try:
                self.imap_client = IMAPClient(host="imap.firstmail.ltd")
                if self.imap_client.login(email_addr, password):
                    self.account_type = "imap"
                    success = True
            except (OSError, ConnectionError, TimeoutError):
                print("IMAP firstmail.ltd: connection failed")

            if not success:
                try:
                    fallback_host = f"imap.{domain}"
                    print(f"Trying fallback IMAP: {fallback_host}")
                    self.imap_client = IMAPClient(host=fallback_host)
                    if self.imap_client.login(email_addr, password):
                        self.account_type = "imap"
                        success = True
                except (OSError, ConnectionError, TimeoutError):
                    print(f"IMAP {domain}: connection failed")

        if success:
            self.last_message_ids = set()
            self.update_status(
                f"Вход выполнен ({self.account_type.upper()}). Получаю письма..."
            )
            self.refresh_inbox_thread(show_loading=True)
        else:
            self.update_status("Ошибка входа (API и IMAP недоступны)")
            self.current_token = None
            self.imap_client = None

    def on_manual_refresh(self):
        """Ручное обновление писем."""
        self.update_status("Обновление писем...")
        threading.Thread(
            target=lambda: self.refresh_inbox_thread(show_loading=True), daemon=True
        ).start()

    def start_auto_refresh(self):
        """Запускает таймер автообновления."""
        if self.stop_threads:
            return

        should_refresh = (self.account_type == "api" and self.current_token) or (
            self.account_type == "imap" and self.imap_client
        )
        if should_refresh and not self.is_refreshing:
            threading.Thread(target=self.refresh_inbox_thread, daemon=True).start()

        self.root.after(self.refresh_interval_ms, self.start_auto_refresh)

    def refresh_inbox_thread(self, show_loading=False):
        """Поток обновления писем."""
        if self.is_refreshing:
            return
        if self.account_type == "api" and not self.current_token:
            return
        if self.account_type == "imap" and not self.imap_client:
            return

        self.is_refreshing = True
        if show_loading:
            self.root.after(0, self.show_inbox_loading_state)
            self.root.after(0, self.show_loading_messages_text)
        try:
            messages = []
            should_update_ui = False
            if self.account_type == "api":
                headers = {"Authorization": f"Bearer {self.current_token}"}
                res = self._make_request(
                    "get", f"{API_URL}/messages", retry_auth=True, headers=headers
                )
                if res is None:
                    self.root.after(
                        0,
                        lambda: self.update_status(
                            "Сетевая ошибка, переподключение..."
                        ),
                    )
                elif res.status_code == 200:
                    messages = res.json()["hydra:member"]
                    should_update_ui = True
                elif res.status_code == 401:
                    self.root.after(
                        0,
                        lambda: self.update_status(
                            "Сессия истекла, переподключение..."
                        ),
                    )
                    self._try_reauth()
                else:
                    self.root.after(
                        0,
                        lambda: self.update_status(
                            f"Ошибка загрузки писем: {res.status_code}"
                        ),
                    )
            elif self.account_type == "imap":
                try:
                    messages = self.imap_client.get_messages(limit=20)
                    should_update_ui = True
                except Exception as imap_err:
                    print(f"IMAP error: {imap_err}")
                    self.root.after(
                        0,
                        lambda: self.update_status(
                            "IMAP соединение потеряно, переподключение..."
                        ),
                    )
                    self._try_reauth()

            if should_update_ui:
                self.root.after(0, lambda msgs=messages: self._update_inbox_ui(msgs))
        except Exception as e:
            print(f"Background update error: {e}")
        finally:
            self.is_refreshing = False

    def _update_inbox_ui(self, messages):
        """Обновление таблицы писем."""
        selected = self.tree.selection()
        selected_id = None
        if selected:
            values = self.tree.item(selected[0]).get("values", [])
            if len(values) >= 4:
                selected_id = values[3]

        for item in self.tree.get_children():
            self.tree.delete(item)

        seen_ids = set()
        new_selection = None

        for msg in messages:
            sender = msg.get("from", {}).get("address", "Неизвестно")
            subject = msg.get("subject") or "(без темы)"
            date_raw = msg.get("createdAt") or ""
            msg_id = msg.get("id")
            try:
                dt = datetime.fromisoformat(str(date_raw).replace("Z", "+00:00"))
                date_str = dt.strftime("%H:%M:%S")
            except Exception:
                date_str = date_raw

            item_id = self.tree.insert(
                "", 0, values=(sender, subject, date_str, msg_id)
            )
            seen_ids.add(msg_id)

            if selected_id and msg_id == selected_id:
                new_selection = item_id

        if new_selection:
            self.tree.selection_set(new_selection)
            self.tree.see(new_selection)

        if not messages and not selected_id:
            self.msg_text.delete(1.0, tk.END)
            self.msg_text.insert(tk.END, "Нет новых писем.")

        new_ids = [mid for mid in seen_ids if mid and mid not in self.last_message_ids]
        if self.last_message_ids and new_ids:
            self.play_notification_sound(len(new_ids))
        self.last_message_ids = seen_ids

        self.status_var.set(
            f"Обновлено: {datetime.now().strftime('%H:%M:%S')} | Писем: {len(messages)}"
        )

    def show_inbox_loading_state(self):
        """Показываем индикатор загрузки."""
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.tree.insert("", 0, values=("Загрузка писем...", "", "", "loading"))
        except Exception:
            pass

    def show_loading_messages_text(self):
        """Показываем текст загрузки."""
        try:
            if self.tree.selection():
                return
            self.btn_copy_code.pack_forget()
            self.msg_text.delete(1.0, tk.END)
            self.msg_text.insert(tk.END, "Загрузка сообщений...")
        except Exception:
            pass

    def on_message_select(self, event):
        """Выбор письма."""
        selected_item = self.tree.selection()
        if not selected_item:
            return

        item = self.tree.item(selected_item)
        values = item.get("values", [])
        if len(values) < 4:
            return
        msg_id = values[3]
        sender = values[0]
        subject = values[1]

        self.btn_copy_code.pack_forget()
        self.msg_text.delete(1.0, tk.END)
        self.msg_text.insert(tk.END, "Загрузка...")

        threading.Thread(
            target=self.load_message_thread, args=(msg_id, sender, subject), daemon=True
        ).start()

    def load_message_thread(self, msg_id, sender=None, subject=None):
        """Поток загрузки письма."""
        if self.account_type == "api" and not self.current_token:
            return
        if self.account_type == "imap" and not self.imap_client:
            return

        try:
            if self.account_type == "api":
                headers = {"Authorization": f"Bearer {self.current_token}"}
                res = self._make_request(
                    "get",
                    f"{API_URL}/messages/{msg_id}",
                    retry_auth=True,
                    headers=headers,
                )
                if res is None:
                    self.root.after(
                        0,
                        lambda: self.msg_text.insert(
                            tk.END, "\nСетевая ошибка при загрузке письма"
                        ),
                    )
                elif res.status_code == 200:
                    data = res.json()
                    text = (
                        data.get("text")
                        or data.get("html")
                        or "Нет текстового содержимого"
                    )
                    self.root.after(0, lambda: self._show_message_content(data, text))
                elif res.status_code == 401:
                    self.root.after(
                        0,
                        lambda: self.msg_text.insert(
                            tk.END, "\nСессия истекла, переподключение..."
                        ),
                    )
                    self._try_reauth()
                else:
                    self.root.after(
                        0,
                        lambda: self.msg_text.insert(
                            tk.END, f"\nОшибка загрузки письма: {res.status_code}"
                        ),
                    )
            elif self.account_type == "imap":
                try:
                    text = self.imap_client.get_message_content(msg_id)
                    data = {
                        "from": {"address": sender or "IMAP Sender"},
                        "subject": subject or "IMAP Message",
                    }
                    self.root.after(
                        0, lambda: self._show_message_content(data, text, is_imap=True)
                    )
                except Exception as imap_err:
                    print(f"IMAP error loading message: {imap_err}")
                    self.root.after(
                        0,
                        lambda: self.msg_text.insert(
                            tk.END, "\nIMAP соединение потеряно, переподключение..."
                        ),
                    )
                    self._try_reauth()
        except Exception as e:
            self.root.after(0, lambda: self.msg_text.insert(tk.END, f"\nError: {e}"))

    def _show_message_content(self, data, text, is_imap=False):
        """Отображение содержимого письма."""
        self.btn_copy_code.pack_forget()
        self.msg_text.delete(1.0, tk.END)

        sender = data.get("from", {}).get("address", "Неизвестно")
        subject = data.get("subject", "(без темы)")
        self.msg_text.insert(tk.END, f"От: {sender}\n")
        self.msg_text.insert(tk.END, f"Тема: {subject}\n")
        self.msg_text.insert(tk.END, "\u2500" * 50 + "\n\n")
        self.msg_text.insert(tk.END, text)

        match = re.search(r"\b(\d{6})\b", text)
        if match:
            code = match.group(1)
            self.btn_copy_code.config(
                text=f"Скопировать код: {code}",
                command=lambda: self.copy_code_to_clipboard(code),
            )
            self.btn_copy_code.pack(
                before=self.msg_text.master, fill=tk.X,
                padx=self.PAD_X, pady=4
            )

    def show_context_menu(self, event):
        """Показ контекстного меню."""
        try:
            self.acc_listbox.selection_clear(0, tk.END)
            self.acc_listbox.selection_set(self.acc_listbox.nearest(event.y))
            self.acc_listbox.activate(self.acc_listbox.nearest(event.y))
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()

    def set_account_status(self, status):
        """Установка статуса аккаунта."""
        selection = self.acc_listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        if idx < len(self.accounts_data):
            self.accounts_data[idx]["status"] = status
            self.update_listbox_colors()
            self.save_accounts_to_file()
            self.update_status(f"Статус обновлен: {status}")
